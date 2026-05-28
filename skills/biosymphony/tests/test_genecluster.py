from __future__ import annotations

import json
import shutil
import sqlite3
import subprocess
import tempfile
import csv
from contextlib import closing
import unittest
from pathlib import Path
from unittest.mock import patch

import sys

ROOT = Path(__file__).resolve().parents[3]
SKILL_ROOT = ROOT / "skills" / "biosymphony"
SCRIPTS = ROOT / "skills" / "biosymphony" / "scripts"
REMOTE = ROOT / "skills" / "biosymphony" / "remote"
EXAMPLE = ROOT / "skills" / "biosymphony" / "examples" / "genecluster-coptis-bia-public-v0"

sys.path.insert(0, str(SCRIPTS))
sys.path.insert(0, str(REMOTE))

import genecluster_claim_audit  # noqa: E402
import genecluster_anchor_map  # noqa: E402
import genecluster_annotation_scout  # noqa: E402
import genecluster_atlas_contracts  # noqa: E402
import genecluster_atlas_normalizers  # noqa: E402
import genecluster_data_materialization  # noqa: E402
import genecluster_db_bootstrap  # noqa: E402
import genecluster_dossier_skeleton  # noqa: E402
import genecluster_excel_intake  # noqa: E402
import genecluster_input_audit  # noqa: E402
import genecluster_issue_dry_run  # noqa: E402
import genecluster_launch_bundle  # noqa: E402
import genecluster_neighborhood_extract  # noqa: E402
import genecluster_neighborhood_score  # noqa: E402
import genecluster_orthology_anchor  # noqa: E402
import genecluster_preflight  # noqa: E402
import genecluster_route_audit  # noqa: E402
import genecluster_reference_import  # noqa: E402
import genecluster_remote_runner  # noqa: E402
import genecluster_runpod_rest_launch  # noqa: E402
import build_runpod_dockerstart  # noqa: E402
import genecluster_stage_contract  # noqa: E402
import genecluster_sra_runinfo  # noqa: E402
import genecluster_source_scout  # noqa: E402
import genecluster_target_db_builder  # noqa: E402
import genecluster_review_surface  # noqa: E402
import biosymphony_public_skill_audit  # noqa: E402
import symphony_orchestration_preflight  # noqa: E402


class GeneClusterPreflightTests(unittest.TestCase):
    def test_public_skill_audit_has_no_private_public_path_errors(self) -> None:
        result = biosymphony_public_skill_audit.audit_skill(SKILL_ROOT)
        self.assertEqual([], result["errors"])

    def test_public_skill_audit_blocks_private_tokens_in_public_paths(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp) / "skill"
            (root / "references").mkdir(parents=True)
            private_home = "/" + "Users" + "/example-operator/autonomy"
            private_slug = "demo" + "_" + "1"
            (root / "references" / "public.md").write_text(
                f"Use {private_home} for this {private_slug} run.\n",
                encoding="utf-8",
            )
            result = biosymphony_public_skill_audit.audit_skill(root)
        rule_ids = {row["rule_id"] for row in result["errors"]}
        self.assertIn("operator_home_path", rule_ids)
        self.assertIn("private_demo_slug", rule_ids)

    def test_public_skill_audit_allows_internal_runbooks(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp) / "skill"
            internal = root / "references" / "internal"
            internal.mkdir(parents=True)
            private_slug = "demo" + "_" + "1"
            (internal / "runbook.md").write_text(
                f"Private runbook path: /opt/autonomy {private_slug}.\n",
                encoding="utf-8",
            )
            result = biosymphony_public_skill_audit.audit_skill(root)
        self.assertEqual([], result["errors"])

    def test_example_ledgers_pass(self) -> None:
        checks = [
            genecluster_preflight.validate_campaign_manifest(EXAMPLE / "campaign-manifest.json"),
            genecluster_preflight.validate_project_goals(EXAMPLE / "project-goals.yaml"),
            genecluster_preflight.validate_pathway_steps(EXAMPLE / "pathway-steps.tsv"),
            genecluster_preflight.validate_data_ledger(EXAMPLE / "data-ledger.tsv"),
            genecluster_preflight.validate_query_ledger(EXAMPLE / "query-ledger.tsv"),
            genecluster_preflight.validate_resource_ledger(EXAMPLE / "resource-ledger.tsv"),
            genecluster_preflight.validate_database_ledger(EXAMPLE / "database-ledger.tsv", repo_root=ROOT),
            genecluster_preflight.validate_cache_ledger(EXAMPLE / "cache-ledger.tsv", repo_root=ROOT),
        ]
        errors = [error for result in checks for error in result["errors"]]
        self.assertEqual([], errors)

    def test_local_artifact_scan_allows_tiny_public_fixtures(self) -> None:
        result = genecluster_preflight.scan_local_artifacts(ROOT)
        self.assertEqual([], result["errors"])

    def test_local_artifact_scan_blocks_unapproved_sequence_files(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            (root / "scratch.faa").write_text(">seq\nMA\n", encoding="utf-8")
            result = genecluster_preflight.scan_local_artifacts(root)
        self.assertFalse(result["ok"])
        self.assertTrue(any("scratch.faa" in error for error in result["errors"]))

    def test_campaign_rejects_local_large_downloads(self) -> None:
        data = json.loads((EXAMPLE / "campaign-manifest.json").read_text(encoding="utf-8"))
        data["execution"]["large_local_downloads"] = True
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "campaign.json"
            path.write_text(json.dumps(data), encoding="utf-8")
            result = genecluster_preflight.validate_campaign_manifest(path)
        self.assertFalse(result["ok"])
        self.assertIn("execution.large_local_downloads must be false", result["errors"])

    def test_data_ledger_rejects_local_remote_path(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "data-ledger.tsv"
            path.write_text(
                "dataset_id\taccession\trun_id\tdata_role\tsample_type\torganism\tbioproject\ttechnology\texpected_size\tsource_url\tremote_path\tchecksum_status\n"
                "bad\tSRX1\tSRR1\ttranscriptome\tleaf\tCoptis chinensis\tPRJ\tIllumina\t1 Gb\thttps://example.org\t./raw.fastq.gz\tpending\n",
                encoding="utf-8",
            )
            result = genecluster_preflight.validate_data_ledger(path)
        self.assertFalse(result["ok"])
        self.assertTrue(any("remote_path must point to remote storage" in error for error in result["errors"]))

    def test_dossier_skeleton_passes_manifest_preflight(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "dossier"
            manifest = genecluster_dossier_skeleton.build_dossier(
                EXAMPLE / "campaign-manifest.json",
                EXAMPLE / "fixtures" / "candidate_hits.tsv",
                out,
            )
            result = genecluster_preflight.validate_dossier_manifest(manifest, repo_root=ROOT)
        self.assertEqual([], result["errors"])

    def test_dossier_skeleton_data_artifacts_pass_table_preflight(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "dossier"
            genecluster_dossier_skeleton.build_dossier(
                EXAMPLE / "campaign-manifest.json",
                EXAMPLE / "fixtures" / "candidate_hits.tsv",
                out,
            )
            checks = [
                genecluster_preflight.validate_candidate_hits(out / "data" / "candidate_hits.tsv"),
                genecluster_preflight.validate_candidate_ranking(out / "data" / "candidate-ranking.tsv"),
                genecluster_preflight.validate_cluster_neighborhoods(out / "data" / "cluster_neighborhoods.tsv"),
                genecluster_preflight.validate_jsonl(
                    out / "data" / "evidence.jsonl",
                    required_keys={"claim_id", "subject_id", "evidence_class", "source_artifact", "confidence", "review_status"},
                    label="evidence.jsonl",
                ),
                genecluster_preflight.validate_claim_audit_jsonl(out / "data" / "claim-audit.jsonl"),
                genecluster_preflight.validate_claim_ledger(out / "claim-ledger.md"),
                genecluster_preflight.validate_database_ledger(out / "data" / "database-ledger.tsv", repo_root=ROOT),
                genecluster_preflight.validate_cache_ledger(out / "data" / "cache-ledger.tsv", repo_root=ROOT),
            ]
        errors = [error for result in checks for error in result["errors"]]
        self.assertEqual([], errors)

    def test_dossier_skeleton_emits_interoperable_sidecars(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "dossier"
            genecluster_dossier_skeleton.build_dossier(
                EXAMPLE / "campaign-manifest.json",
                EXAMPLE / "fixtures" / "candidate_hits.tsv",
                out,
            )
            datapackage = json.loads((out / "datapackage.json").read_text(encoding="utf-8"))
            crate = json.loads((out / "ro-crate-metadata.json").read_text(encoding="utf-8"))
            validation_report = json.loads((out / "validation-report.json").read_text(encoding="utf-8"))

        self.assertEqual("data-package", datapackage["profile"])
        resource_paths = {resource["path"] for resource in datapackage["resources"]}
        self.assertIn("data/candidate_hits.tsv", resource_paths)
        self.assertIn("data/provenance.jsonl", resource_paths)
        for resource_path in resource_paths:
            self.assertFalse(Path(resource_path).is_absolute())

        graph_ids = {entity["@id"] for entity in crate["@graph"] if "@id" in entity}
        self.assertIn("./", graph_ids)
        self.assertIn("#create-dossier-skeleton", graph_ids)
        self.assertIn("datapackage.json", graph_ids)
        self.assertIn("dossier-manifest.json", graph_ids)
        self.assertEqual("passed", validation_report["status"])
        self.assertTrue(validation_report["checks"])

    def test_dossier_claim_ledger_drives_review_surface(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            dossier = tmp_path / "dossier"
            review = tmp_path / "review"
            genecluster_dossier_skeleton.build_dossier(
                EXAMPLE / "campaign-manifest.json",
                EXAMPLE / "fixtures" / "candidate_hits.tsv",
                dossier,
            )
            build_result = genecluster_review_surface.build_review_surface(
                review,
                review_id="dossier-claim-review",
                claim_ledger=dossier / "claim-ledger.tsv",
            )
            manifest = Path(build_result["manifest"])
            manifest_data = json.loads(manifest.read_text(encoding="utf-8"))
            validate_result = genecluster_atlas_contracts.validate_review_surface_manifest(manifest)
            review_index = (review / "index.html").read_text(encoding="utf-8")

        self.assertEqual([], validate_result["errors"])
        self.assertEqual(3, len(manifest_data["claims"]))
        self.assertIn("Claim Summary", review_index)
        self.assertIn("claim-ledger-only", review_index)

    def test_dossier_sidecars_reject_raw_or_absolute_paths(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "dossier"
            manifest = genecluster_dossier_skeleton.build_dossier(
                EXAMPLE / "campaign-manifest.json",
                EXAMPLE / "fixtures" / "candidate_hits.tsv",
                out,
            )
            datapackage_path = out / "datapackage.json"
            datapackage = json.loads(datapackage_path.read_text(encoding="utf-8"))
            datapackage["resources"].append(
                {
                    "name": "raw-fastq",
                    "path": "/private/raw.fastq.gz",
                    "hash": "0" * 64,
                }
            )
            datapackage_path.write_text(json.dumps(datapackage), encoding="utf-8")
            result = genecluster_preflight.validate_dossier_manifest(manifest, repo_root=ROOT)
        self.assertFalse(result["ok"])
        self.assertTrue(any("resource path must be relative" in error for error in result["errors"]))

    def test_issue_dry_run_outputs_valid_contracts(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "issues"
            issues = genecluster_issue_dry_run.build_issues(EXAMPLE / "campaign-manifest.json", "GENECLUSTER-MIT")
            out.mkdir()
            for iid, body in issues.items():
                (out / f"{iid}.md").write_text(body, encoding="utf-8")
                result = genecluster_issue_dry_run.validate_issue(body)
                self.assertEqual([], result["errors"])

    def test_launch_bundles_validate_for_supported_providers(self) -> None:
        cases = [
            ("local_lite", "smoke", "", False),
            ("local_lite", "next_experiment_design", "", False),
            ("local_full", "candidate_search", "/tmp/genecluster-local-full-test", True),
            ("local_full", "full_public_mining", "/tmp/genecluster-local-full-test-public-mining", True),
            ("runpod_pod", "full_campaign", "", False),
            ("runpod_pod", "full_public_mining", "", False),
            ("ssh_hpc", "synteny", "/remote/genecluster-runs/test", False),
            ("cloud_vm", "genome_context", "/mnt/genecluster-runs/test", False),
        ]
        with tempfile.TemporaryDirectory() as tmp:
            for provider, scope, workdir, allow_local_full in cases:
                out = Path(tmp) / f"{provider}-{scope}"
                manifest = genecluster_launch_bundle.build_launch_bundle(
                    campaign_path=EXAMPLE / "campaign-manifest.json",
                    out=out,
                    provider_class=provider,
                    run_scope=scope,
                    repo_root=ROOT,
                    heavy_workdir=workdir,
                    run_id=f"{provider}-{scope}",
                    allow_local_full=allow_local_full,
                )
                result = genecluster_preflight.validate_launch_manifest(manifest, repo_root=ROOT)
                self.assertEqual([], result["errors"])

    def test_launch_scope_alias_validates(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            manifest = genecluster_launch_bundle.build_launch_bundle(
                campaign_path=EXAMPLE / "campaign-manifest.json",
                out=Path(tmp) / "bundle",
                provider_class="runpod-pod",
                run_scope="full-public-mining",
                repo_root=ROOT,
                run_id="alias-full-public-mining",
            )
            data = json.loads(manifest.read_text(encoding="utf-8"))
            result = genecluster_preflight.validate_launch_manifest(manifest, repo_root=ROOT)
        self.assertEqual("runpod_pod", data["provider_class"])
        self.assertEqual("full_public_mining", data["run_scope"])
        self.assertEqual([], result["errors"])

    def test_excel_intake_generates_private_campaign_bundle(self) -> None:
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            workbook = tmp_path / "demo.xlsx"
            wb = Workbook()
            ws = wb.active
            ws["B6"] = "Links to Coptis sequence data"
            ws["B7"] = "transcriptome 1"
            ws["C7"] = "https://www.ncbi.nlm.nih.gov/sra/SRX9153204[accn]"
            ws["B8"] = "Genome"
            ws["C8"] = "https://www.ncbi.nlm.nih.gov/sra/SRX9153201[accn]"
            ws["E8"] = "Find all clusters for select queries"
            ws["B26"] = "Protein queries"
            ws["B27"] = "#"
            ws["C27"] = "Name"
            ws["D27"] = "Description"
            ws["E27"] = "Sequence"
            ws["B28"] = 1
            ws["C28"] = "C.roseusSTR1_T"
            ws["D28"] = "Catharanthus roseus strictosidine synthase"
            ws["E28"] = "MSPILKKIFIESPSYAPNAFTFDSTDKG"
            ws["B29"] = 2
            ws["C29"] = "secologanin transporter, BIA_transporter,"
            ws["E29"] = "<- you will have to find."
            wb.save(workbook)

            out = tmp_path / "intake"
            genecluster_excel_intake.intake_workbook(workbook, out, "genecluster-private-demo-test")
            checks = [
                genecluster_preflight.validate_campaign_manifest(out / "campaign-manifest.json"),
                genecluster_preflight.validate_project_goals(out / "project-goals.yaml"),
                genecluster_preflight.validate_pathway_steps(out / "pathway-steps.tsv"),
                genecluster_preflight.validate_data_ledger(out / "data-ledger.tsv"),
                genecluster_preflight.validate_query_ledger(out / "query-ledger.tsv"),
                genecluster_preflight.validate_resource_ledger(out / "resource-ledger.tsv"),
                genecluster_preflight.validate_database_ledger(out / "database-ledger.tsv", repo_root=ROOT),
                genecluster_preflight.validate_cache_ledger(out / "cache-ledger.tsv", repo_root=ROOT),
            ]
            errors = [error for result in checks for error in result["errors"]]
            self.assertEqual([], errors)
            self.assertTrue((out / "query-sequences.faa").exists())

            bundle = genecluster_launch_bundle.build_launch_bundle(
                campaign_path=out / "campaign-manifest.json",
                out=tmp_path / "bundle",
                provider_class="runpod_pod",
                run_scope="full_campaign_24h",
                repo_root=ROOT,
                run_id="excel-private-demo",
            )
            plan = json.loads((bundle.parent / "query-resolution-plan.json").read_text(encoding="utf-8"))
            actions = {record["resolution_action"] for record in plan["records"]}
            self.assertIn("use_embedded_query_fasta", actions)
            self.assertTrue((bundle.parent / "inputs" / "query-sequences.faa").exists())

    def test_local_lite_rejects_heavy_scope(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            with self.assertRaises(ValueError):
                genecluster_launch_bundle.build_launch_bundle(
                    campaign_path=EXAMPLE / "campaign-manifest.json",
                    out=Path(tmp) / "bundle",
                    provider_class="local_lite",
                    run_scope="candidate_search",
                    repo_root=ROOT,
                    run_id="bad-local-lite",
                )

    def test_local_full_rejects_repo_workdir(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            with self.assertRaises(ValueError):
                genecluster_launch_bundle.build_launch_bundle(
                    campaign_path=EXAMPLE / "campaign-manifest.json",
                    out=Path(tmp) / "bundle",
                    provider_class="local_full",
                    run_scope="candidate_search",
                    repo_root=ROOT,
                    heavy_workdir=str(ROOT / "genecluster-runs" / "bad"),
                    run_id="bad-local-full",
                    allow_local_full=True,
                )

    def test_runpod_bundle_reports_missing_credential_names_only(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            with patch.dict("os.environ", {"RUNPOD_API_KEY": ""}, clear=False):
                manifest = genecluster_launch_bundle.build_launch_bundle(
                    campaign_path=EXAMPLE / "campaign-manifest.json",
                    out=Path(tmp) / "bundle",
                    provider_class="runpod_pod",
                    run_scope="smoke",
                    repo_root=ROOT,
                    run_id="runpod-smoke",
                )
            data = json.loads(manifest.read_text(encoding="utf-8"))
        self.assertIn("RUNPOD_API_KEY", data["missing_credentials"])
        self.assertTrue(all("=" not in item for item in data["missing_credentials"]))

    def test_launch_ready_rejects_missing_runpod_credentials(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            with patch.dict("os.environ", {"RUNPOD_API_KEY": ""}, clear=False):
                manifest = genecluster_launch_bundle.build_launch_bundle(
                    campaign_path=EXAMPLE / "campaign-manifest.json",
                    out=Path(tmp) / "bundle",
                    provider_class="runpod_pod",
                    run_scope="smoke",
                    repo_root=ROOT,
                    run_id="runpod-launch-ready",
                )
            result = genecluster_preflight.validate_launch_manifest(
                manifest,
                repo_root=ROOT,
                launch_ready=True,
            )
        self.assertFalse(result["ok"])
        self.assertTrue(any("launch-ready validation requires" in error for error in result["errors"]))

    def test_runpod_launch_ready_rejects_placeholder_image_and_unresolved_volume(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            with patch.dict(
                "os.environ",
                {
                    "RUNPOD_API_KEY": "present",
                    "GENECLUSTER_RUNPOD_NETWORK_VOLUME_ID": "",
                    "GENECLUSTER_RUNPOD_DATACENTER": "",
                },
                clear=False,
            ):
                manifest = genecluster_launch_bundle.build_launch_bundle(
                    campaign_path=EXAMPLE / "campaign-manifest.json",
                    out=Path(tmp) / "bundle",
                    provider_class="runpod_pod",
                    run_scope="full_campaign",
                    repo_root=ROOT,
                    run_id="runpod-launch-ready-placeholder",
                )
            result = genecluster_preflight.validate_launch_manifest(
                manifest,
                repo_root=ROOT,
                launch_ready=True,
            )
        self.assertFalse(result["ok"])
        self.assertTrue(any("non-placeholder runner image" in error or "image is a placeholder" in error for error in result["errors"]))
        self.assertTrue(any("network_volume_id" in error for error in result["errors"]))

    def test_execution_ready_requires_db_cache_search_contracts(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            manifest = genecluster_launch_bundle.build_launch_bundle(
                campaign_path=EXAMPLE / "campaign-manifest.json",
                out=Path(tmp) / "bundle",
                provider_class="runpod_pod",
                run_scope="full_campaign",
                repo_root=ROOT,
                run_id="runpod-execution-contract",
                image="ghcr.io/example/genecluster@sha256:" + "a" * 64,
            )
            data = json.loads(manifest.read_text(encoding="utf-8"))
            data.pop("database_ledger")
            data.pop("launch_payload_sha256", None)
            manifest.write_text(json.dumps(data), encoding="utf-8")
            result = genecluster_preflight.validate_launch_manifest(
                manifest,
                repo_root=ROOT,
                execution_ready=True,
            )
        self.assertFalse(result["ok"])
        self.assertTrue(any("database_ledger" in error for error in result["errors"]))

    def test_launch_bundle_emits_provider_payloads_and_plans(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            manifest = genecluster_launch_bundle.build_launch_bundle(
                campaign_path=EXAMPLE / "campaign-manifest.json",
                out=Path(tmp) / "bundle",
                provider_class="runpod_pod",
                run_scope="full_campaign",
                repo_root=ROOT,
                run_id="payloads-and-plans",
            )
            bundle = manifest.parent
            checks = [
                genecluster_preflight.validate_db_bootstrap_plan(bundle / "db-bootstrap-plan.json"),
                genecluster_preflight.validate_data_materialization_plan(bundle / "data-materialization-plan.json"),
                genecluster_preflight.validate_reference_import_plan(bundle / "reference-import-plan.json"),
                genecluster_preflight.validate_anchor_map_plan(bundle / "anchor-map-plan.json"),
                genecluster_preflight.validate_neighborhood_extract_plan(bundle / "neighborhood-extract-plan.json"),
                genecluster_preflight.validate_query_resolution_plan(bundle / "query-resolution-plan.json"),
                genecluster_preflight.validate_decoy_plan(bundle / "decoy-plan.json"),
                genecluster_preflight.validate_run_economics(bundle / "run-economics.json"),
                genecluster_preflight.validate_search_plan(bundle / "search-plan.json"),
                genecluster_preflight.validate_tool_requirements(bundle / "tool-requirements.json"),
                genecluster_preflight.validate_provider_payload(bundle / "provider" / "runpod-pod.json", repo_root=ROOT),
                genecluster_preflight.validate_provider_payload(bundle / "provider" / "local-full.sh", repo_root=ROOT),
                genecluster_preflight.validate_provider_payload(bundle / "provider" / "ssh-hpc.sh", repo_root=ROOT),
                genecluster_preflight.validate_provider_payload(bundle / "provider" / "cloud-vm.sh", repo_root=ROOT),
            ]
        errors = [error for result in checks for error in result["errors"]]
        self.assertEqual([], errors)

    def test_launch_bundle_emits_stage_contract(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            manifest = genecluster_launch_bundle.build_launch_bundle(
                campaign_path=EXAMPLE / "campaign-manifest.json",
                out=Path(tmp) / "bundle",
                provider_class="runpod_pod",
                run_scope="candidate_search",
                repo_root=ROOT,
                run_id="stage-contract",
            )
            bundle = manifest.parent
            launch = json.loads(manifest.read_text(encoding="utf-8"))
            result = genecluster_stage_contract.validate_stage_contract(bundle / "stage-contract.json")
            stage_data = json.loads((bundle / "stage-contract.json").read_text(encoding="utf-8"))
        self.assertEqual([], result["errors"])
        self.assertIn("stage_contract", launch["ledger_hashes"])
        self.assertEqual("stage-contract.json", launch["stage_contract"])
        self.assertTrue(any("genecluster_stage_contract.py" in command for command in launch["validation_commands"]))
        stage_ids = {row["stage_id"] for row in stage_data["stages"]}
        self.assertIn("candidate_search", stage_ids)
        self.assertEqual(10, stage_data["heartbeat_interval_minutes"])

    def test_stage_contract_requires_fail_closed_tool_proofs(self) -> None:
        contract = {
            "schema_version": 1,
            "run_id": "tool-proof-test",
            "provider_class": "runpod_pod",
            "run_scope": "candidate_search",
            "progress_ledger": "/workspace/genecluster/runs/tool-proof-test/summary/stage-progress.jsonl",
            "heartbeat_interval_minutes": 10,
            "stale_after_minutes": 30,
            "stages": [
                {
                    "stage_id": "predict_orfs",
                    "run_flag": "--data-materialization",
                    "entrypoint": "TransDecoder.Predict --single_best_only -t transcripts.fa",
                    "expected_outputs": ["proteins.faa"],
                    "done_marker": "/workspace/genecluster/runs/tool-proof-test/summary/proteins.faa",
                    "timeout_minutes": 30,
                    "resume_strategy": "idempotent rerun after output check",
                    "failure_policy": "fail-closed before downstream candidate search",
                }
            ],
            "watcher": {
                "required_for_runtime_hours_over": 2,
                "poll_interval_minutes": 10,
                "false_assumptions_to_avoid": ["do not assume stale progress means provider capacity"],
            },
            "acceptance": {
                "final_success_requires": ["terminal progress row"],
                "partial_verdict_allowed": True,
                "partial_verdict_requires": ["failed stage id"],
            },
        }
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "stage-contract.json"
            path.write_text(json.dumps(contract), encoding="utf-8")
            result = genecluster_stage_contract.validate_stage_contract(path)
        self.assertFalse(result["ok"])
        self.assertTrue(any("TransDecoder.Predict" in error for error in result["errors"]))

    def test_stage_contract_rejects_warning_only_required_tools(self) -> None:
        contract = {
            "schema_version": 1,
            "run_id": "warning-tool-test",
            "provider_class": "runpod_pod",
            "run_scope": "candidate_search",
            "progress_ledger": "/workspace/genecluster/runs/warning-tool-test/summary/stage-progress.jsonl",
            "heartbeat_interval_minutes": 10,
            "stale_after_minutes": 30,
            "stages": [
                {
                    "stage_id": "predict_orfs",
                    "run_flag": "--data-materialization",
                    "entrypoint": "remote/genecluster_remote_runner.py",
                    "expected_outputs": ["proteins.faa"],
                    "done_marker": "/workspace/genecluster/runs/warning-tool-test/summary/proteins.faa",
                    "timeout_minutes": 30,
                    "resume_strategy": "idempotent rerun after output check",
                    "failure_policy": "fail-closed before downstream candidate search",
                    "required_tools": [
                        {
                            "name": "TransDecoder.Predict",
                            "executable": "TransDecoder.Predict",
                            "proof_command": "TransDecoder.Predict --help",
                            "fail_closed": False,
                        }
                    ],
                }
            ],
            "watcher": {
                "required_for_runtime_hours_over": 2,
                "poll_interval_minutes": 10,
                "false_assumptions_to_avoid": ["do not assume stale progress means provider capacity"],
            },
            "acceptance": {
                "final_success_requires": ["terminal progress row"],
                "partial_verdict_allowed": True,
                "partial_verdict_requires": ["failed stage id"],
            },
        }
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "stage-contract.json"
            path.write_text(json.dumps(contract), encoding="utf-8")
            result = genecluster_stage_contract.validate_stage_contract(path)
        self.assertFalse(result["ok"])
        self.assertTrue(any("fail_closed=true" in error for error in result["errors"]))

    def test_full_bundle_stage_contract_proves_transcript_first_tools(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            manifest = genecluster_launch_bundle.build_launch_bundle(
                campaign_path=EXAMPLE / "campaign-manifest.json",
                out=Path(tmp) / "bundle",
                provider_class="runpod_pod",
                run_scope="full_campaign_24h",
                repo_root=ROOT,
                run_id="stage-tool-proof",
            )
            bundle = manifest.parent
            stage_data = json.loads((bundle / "stage-contract.json").read_text(encoding="utf-8"))
            result = genecluster_stage_contract.validate_stage_contract(bundle / "stage-contract.json")
        tool_names = {tool["name"] for tool in stage_data["required_tools"]}
        self.assertEqual([], result["errors"])
        self.assertIn("hisat2", tool_names)
        self.assertIn("stringtie", tool_names)
        self.assertIn("gffread", tool_names)
        self.assertIn("TransDecoder.LongOrfs", tool_names)
        self.assertIn("TransDecoder.Predict", tool_names)

    def test_runpod_payload_contains_live_run_guardrails(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            manifest = genecluster_launch_bundle.build_launch_bundle(
                campaign_path=EXAMPLE / "campaign-manifest.json",
                out=Path(tmp) / "bundle",
                provider_class="runpod_pod",
                run_scope="full_campaign_24h",
                repo_root=ROOT,
                run_id="runpod-live-guardrails",
                image="ghcr.io/example/genecluster@sha256:" + "a" * 64,
            )
            bundle = manifest.parent
            payload_path = bundle / "provider" / "runpod-pod.json"
            artifact_pull = bundle / "artifact_pull.yaml"
            payload = json.loads(payload_path.read_text(encoding="utf-8"))
            start_script = bundle / payload["start_script"]
            start_script_exists = start_script.exists()
            result = genecluster_preflight.validate_provider_payload(
                payload_path,
                repo_root=ROOT,
                execution_ready=False,
            )
            artifact_pull_result = genecluster_preflight.validate_artifact_pull_manifest(artifact_pull, repo_root=ROOT)
        self.assertEqual([], result["errors"])
        self.assertEqual([], artifact_pull_result["errors"])
        self.assertTrue(start_script_exists)
        self.assertFalse(payload["pod_lifecycle_policy"]["self_stop_on_completion"])
        self.assertTrue(payload["pod_lifecycle_policy"]["operator_side_cleanup_required"])
        self.assertFalse(payload["pod_lifecycle_policy"]["provider_api_key_inside_pod"])
        self.assertGreaterEqual(payload["pod_lifecycle_policy"]["idle_after_completion_seconds"], 60)
        self.assertTrue(payload["pod_lifecycle_policy"]["watch_runtime_uptime_seconds_required"])
        self.assertNotIn("RUNPOD_POD_ID", payload["env_var_names"])
        self.assertNotIn("NCBI_API_KEY", payload["env_var_names"])
        self.assertEqual("runpod_s3_or_configured_summary_endpoint", payload["summary_sync_policy"]["preferred_transport"])
        self.assertEqual("emergency_only", payload["image_policy"]["first_boot_mamba_install"])
        self.assertFalse(payload["image_policy"]["first_boot_install_allowed_for_standard_launch"])
        self.assertEqual("baked_image_required", payload["image_policy"]["tool_install_strategy"])
        self.assertIn("update_blastdb.pl", payload["image_policy"]["required_boot_tools"])

    def test_artifact_pull_manifest_rejects_raw_include(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            manifest = genecluster_launch_bundle.build_launch_bundle(
                campaign_path=EXAMPLE / "campaign-manifest.json",
                out=Path(tmp) / "bundle",
                provider_class="runpod_pod",
                run_scope="candidate_search",
                repo_root=ROOT,
                run_id="artifact-pull-raw-include",
            )
            artifact_pull = manifest.parent / "artifact_pull.yaml"
            data = json.loads(artifact_pull.read_text(encoding="utf-8"))
            data["include"].append({"path": "raw/sample.fastq.gz", "required": True})
            artifact_pull.write_text(json.dumps(data), encoding="utf-8")
            result = genecluster_preflight.validate_artifact_pull_manifest(artifact_pull, repo_root=ROOT)
        self.assertFalse(result["ok"])
        self.assertTrue(any("raw/large artifact" in error for error in result["errors"]))

    def test_launch_manifest_validates_artifact_pull_manifest(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            manifest = genecluster_launch_bundle.build_launch_bundle(
                campaign_path=EXAMPLE / "campaign-manifest.json",
                out=Path(tmp) / "bundle",
                provider_class="runpod_pod",
                run_scope="candidate_search",
                repo_root=ROOT,
                run_id="launch-validates-artifact-pull",
            )
            artifact_pull = manifest.parent / "artifact_pull.yaml"
            artifact_pull_data = json.loads(artifact_pull.read_text(encoding="utf-8"))
            artifact_pull_data["include"].append({"path": "work/raw-reads.fastq.gz", "required": True})
            artifact_pull.write_text(json.dumps(artifact_pull_data), encoding="utf-8")

            launch_data = json.loads(manifest.read_text(encoding="utf-8"))
            launch_data["ledger_hashes"]["artifact_pull_manifest"] = genecluster_preflight.sha256_file(artifact_pull)
            launch_data.pop("launch_payload_sha256", None)
            manifest.write_text(json.dumps(launch_data), encoding="utf-8")

            result = genecluster_preflight.validate_launch_manifest(manifest, repo_root=ROOT)
        self.assertFalse(result["ok"])
        self.assertTrue(any(error.startswith("artifact_pull_manifest:") for error in result["errors"]))
        self.assertTrue(any("raw/large artifact" in error for error in result["errors"]))

    def test_runpod_payload_blocks_auth_sensitive_registry_without_pull_auth(self) -> None:
        env = {
            "RUNPOD_API_KEY": "present",
            "GENECLUSTER_RUNPOD_NETWORK_VOLUME_ID": "vol-test",
            "GENECLUSTER_RUNPOD_DATACENTER": "US-KS-2",
            "GENECLUSTER_RUNPOD_CONTAINER_REGISTRY_AUTH_ID": "",
            "RUNPOD_CONTAINER_REGISTRY_AUTH_ID": "",
            "GENECLUSTER_CONTAINER_REGISTRY_AUTH_ID": "",
            "GENECLUSTER_RUNPOD_IMAGE_PUBLIC_PULL": "",
            "GENECLUSTER_IMAGE_PUBLIC_PULL": "",
        }
        with tempfile.TemporaryDirectory() as tmp, patch.dict("os.environ", env, clear=False):
            manifest = genecluster_launch_bundle.build_launch_bundle(
                campaign_path=EXAMPLE / "campaign-manifest.json",
                out=Path(tmp) / "bundle",
                provider_class="runpod_pod",
                run_scope="smoke",
                repo_root=ROOT,
                run_id="runpod-registry-auth-blocker",
                image="ghcr.io/example/private-genecluster@sha256:" + "a" * 64,
            )
            payload_path = manifest.parent / "provider" / "runpod-pod.json"
            payload = json.loads(payload_path.read_text(encoding="utf-8"))
            result = genecluster_preflight.validate_provider_payload(
                payload_path,
                repo_root=ROOT,
                execution_ready=True,
            )
        self.assertTrue(payload["registry_auth_policy"]["launch_blocker_if_missing"])
        self.assertFalse(result["ok"])
        self.assertTrue(any("registry likely requires auth" in error for error in result["errors"]))

    def test_runpod_payload_accepts_registry_auth_env_for_private_image(self) -> None:
        env = {
            "RUNPOD_API_KEY": "present",
            "GENECLUSTER_RUNPOD_NETWORK_VOLUME_ID": "vol-test",
            "GENECLUSTER_RUNPOD_DATACENTER": "US-KS-2",
            "GENECLUSTER_RUNPOD_CONTAINER_REGISTRY_AUTH_ID": "cra_test",
            "RUNPOD_CONTAINER_REGISTRY_AUTH_ID": "",
            "GENECLUSTER_CONTAINER_REGISTRY_AUTH_ID": "",
            "GENECLUSTER_RUNPOD_IMAGE_PUBLIC_PULL": "",
            "GENECLUSTER_IMAGE_PUBLIC_PULL": "",
        }
        with tempfile.TemporaryDirectory() as tmp, patch.dict("os.environ", env, clear=False):
            manifest = genecluster_launch_bundle.build_launch_bundle(
                campaign_path=EXAMPLE / "campaign-manifest.json",
                out=Path(tmp) / "bundle",
                provider_class="runpod_pod",
                run_scope="smoke",
                repo_root=ROOT,
                run_id="runpod-registry-auth-ok",
                image="ghcr.io/example/private-genecluster@sha256:" + "a" * 64,
            )
            payload_path = manifest.parent / "provider" / "runpod-pod.json"
            payload = json.loads(payload_path.read_text(encoding="utf-8"))
            result = genecluster_preflight.validate_provider_payload(
                payload_path,
                repo_root=ROOT,
                execution_ready=True,
            )
            auth_env, auth_id = genecluster_runpod_rest_launch.registry_auth_id_for_payload(payload)
        self.assertEqual([], result["errors"])
        self.assertFalse(payload["registry_auth_policy"]["launch_blocker_if_missing"])
        self.assertEqual("GENECLUSTER_RUNPOD_CONTAINER_REGISTRY_AUTH_ID", auth_env)
        self.assertEqual("cra_test", auth_id)

    def test_runpod_launcher_rejects_auth_sensitive_image_before_create(self) -> None:
        provider = {
            "registry_auth_policy": {
                "launch_blocker_if_missing": True,
                "container_registry_auth_id_env_names": ["GENECLUSTER_RUNPOD_CONTAINER_REGISTRY_AUTH_ID"],
                "public_image_assertion_env_names": ["GENECLUSTER_RUNPOD_IMAGE_PUBLIC_PULL"],
                "container_registry_auth_id_present": False,
                "public_image_asserted": False,
            }
        }
        with patch.dict(
            "os.environ",
            {
                "GENECLUSTER_RUNPOD_CONTAINER_REGISTRY_AUTH_ID": "",
                "GENECLUSTER_RUNPOD_IMAGE_PUBLIC_PULL": "",
            },
            clear=False,
        ):
            error = genecluster_runpod_rest_launch.registry_auth_error(provider)
        self.assertIn("image registry likely requires auth", error)

    def test_materialization_plan_distinguishes_transcript_sra_from_genome_defer(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            manifest = genecluster_launch_bundle.build_launch_bundle(
                campaign_path=EXAMPLE / "campaign-manifest.json",
                out=Path(tmp) / "bundle",
                provider_class="runpod_pod",
                run_scope="full_campaign_24h",
                repo_root=ROOT,
                run_id="materialization-contract",
            )
            bundle = manifest.parent
            plan = json.loads((bundle / "data-materialization-plan.json").read_text(encoding="utf-8"))
            target_plan = json.loads((bundle / "target-db-plan.json").read_text(encoding="utf-8"))
            result = genecluster_preflight.validate_data_materialization_plan(bundle / "data-materialization-plan.json")
        self.assertEqual([], result["errors"])
        self.assertEqual(2, plan["summary"]["materializable_raw_sra_source_count"])
        self.assertEqual(1, plan["summary"]["unsupported_raw_sra_source_count"])
        self.assertTrue(any(row["engine"] == "blast" and row["sequence_type"] == "nucleotide" for row in target_plan["index_targets"]))

    def test_superpower_bundle_emits_cross_species_contracts_and_prompt(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            manifest = genecluster_launch_bundle.build_launch_bundle(
                campaign_path=EXAMPLE / "campaign-manifest.json",
                out=Path(tmp) / "bundle",
                provider_class="runpod_pod",
                run_scope="full_campaign_24h",
                repo_root=ROOT,
                run_id="cross-species-superpower",
            )
            bundle = manifest.parent
            launch = json.loads(manifest.read_text(encoding="utf-8"))
            checks = [
                genecluster_preflight.validate_target_db_plan(bundle / "target-db-plan.json", repo_root=ROOT),
                genecluster_preflight.validate_candidate_route_plan(bundle / "candidate-route-plan.json"),
                genecluster_preflight.validate_orthology_anchor_plan(bundle / "orthology-anchor-plan.json"),
                genecluster_preflight.validate_reciprocal_search_plan(bundle / "reciprocal-search-plan.json"),
                genecluster_preflight.validate_pathway_completeness_plan(bundle / "pathway-completeness-plan.json"),
            ]
            prompt = (bundle / "campaign-prompt.md").read_text(encoding="utf-8")
        errors = [error for result in checks for error in result["errors"]]
        self.assertEqual([], errors)
        self.assertIn("target_db_plan", launch["ledger_hashes"])
        self.assertIn("candidate_route_plan", launch["ledger_hashes"])
        self.assertIn("campaign_prompt", launch["ledger_hashes"])
        self.assertIn("canonical proteins from source species A -> target species B", prompt)
        self.assertIn("miniprot", prompt)
        self.assertIn("Candidate route plan", prompt)
        self.assertIn("Input-first rule", prompt)
        self.assertIn("ledgers/data-ledger.tsv", prompt)

    def test_input_audit_surfaces_known_accessions_before_questions(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            manifest = genecluster_launch_bundle.build_launch_bundle(
                campaign_path=EXAMPLE / "campaign-manifest.json",
                out=Path(tmp) / "bundle",
                provider_class="runpod_pod",
                run_scope="full_campaign_24h",
                repo_root=ROOT,
                run_id="input-audit-test",
            )
            audit = genecluster_input_audit.build_audit(manifest)
            prompt = json.loads(manifest.read_text(encoding="utf-8"))["validation_commands"][0]
        accessions = {row.get("accession") for row in audit["known_data_refs"]}
        self.assertIn("SRX9153204", accessions)
        self.assertIn("SRX16999876", accessions)
        self.assertTrue(any("data links/accessions already present" in item for item in audit["input_first_policy"]["do_not_ask_for"]))
        self.assertTrue(audit["input_first_policy"]["read_before_asking"])
        self.assertIn("genecluster_input_audit.py", prompt)
        self.assertIn("--require-known-data", prompt)
        self.assertIn("--interview-mode standard", prompt)

    def test_input_audit_interview_asks_only_unanswered_questions(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            manifest = genecluster_launch_bundle.build_launch_bundle(
                campaign_path=EXAMPLE / "campaign-manifest.json",
                out=Path(tmp) / "bundle",
                provider_class="runpod_pod",
                run_scope="full_campaign_24h",
                repo_root=ROOT,
                run_id="input-interview-test",
            )
            standard = genecluster_input_audit.build_audit(manifest, interview_mode="standard")["intake_interview"]
            quick = genecluster_input_audit.build_audit(manifest, interview_mode="quick")["intake_interview"]
            skipped = genecluster_input_audit.build_audit(manifest, interview_mode="skip")["intake_interview"]
        known_topics = {row["topic"] for row in standard["known"]}
        question_ids = {row["question_id"] for row in standard["questions"]}
        self.assertIn("data_scope", known_topics)
        self.assertIn("query_scope", known_topics)
        self.assertNotIn("data_scope", question_ids)
        self.assertNotIn("query_scope", question_ids)
        self.assertIn("route_blocker_decision", question_ids)
        self.assertLessEqual(quick["counts"]["questions"], standard["counts"]["questions"])
        self.assertEqual([], skipped["questions"])
        self.assertTrue(any(row["topic"] == "operator_interview" for row in skipped["assumptions"]))

    def test_orchestration_preflight_rejects_empty_issue_body_prompt(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            prompt = Path(tmp) / "prompt.md"
            prompt.write_text("<issue_body></issue_body>\n", encoding="utf-8")
            result = symphony_orchestration_preflight.validate_rendered_prompt(prompt)
        self.assertFalse(result["ok"])
        self.assertTrue(any("empty <issue_body>" in error for error in result["errors"]))

    def test_orchestration_preflight_accepts_rendered_genecluster_issue(self) -> None:
        body = next(iter(genecluster_issue_dry_run.build_issues(EXAMPLE / "campaign-manifest.json", "GC").values()))
        with tempfile.TemporaryDirectory() as tmp:
            prompt = Path(tmp) / "prompt.md"
            prompt.write_text(f"<issue_body>\n{body}\n</issue_body>\n", encoding="utf-8")
            result = symphony_orchestration_preflight.validate_rendered_prompt(prompt)
        self.assertEqual([], result["errors"])
        self.assertGreater(result["issue_body_bytes"], 200)

    def test_runpod_payload_preflight_rejects_oversized_payload(self) -> None:
        payload = {"dockerStartCmd": ["bash", "-lc", "x" * (70 * 1024)], "env": {"RUNPOD_API_KEY": "secret"}}
        result = genecluster_runpod_rest_launch.payload_preflight(payload, warn_bytes=50 * 1024, max_bytes=60 * 1024)
        self.assertFalse(result["ok"])
        self.assertTrue(any("too large" in error for error in result["errors"]))
        self.assertTrue(any("secret-like env var" in error for error in result["errors"]))

    def test_runpod_rest_launcher_runtime_env_excludes_secrets(self) -> None:
        provider = {
            "db_cache_root": "/workspace/cache",
            "pod_lifecycle_policy": {"idle_after_completion_seconds": 120},
        }
        with patch.dict(
            "os.environ",
            {
                "RUNPOD_API_KEY": "secret",
                "GITHUB_TOKEN": "secret",
                "NCBI_API_KEY": "secret",
                "GENECLUSTER_FASTQ_THREADS": "8",
            },
            clear=False,
        ):
            env = genecluster_runpod_rest_launch.provider_runtime_env(provider, run_id="runpod-env-test")
        self.assertEqual("runpod-env-test", env["GENECLUSTER_RUN_ID"])
        self.assertEqual("8", env["GENECLUSTER_FASTQ_THREADS"])
        self.assertEqual("120", env["GENECLUSTER_RUNPOD_IDLE_SECONDS"])
        self.assertNotIn("RUNPOD_API_KEY", env)
        self.assertNotIn("GITHUB_TOKEN", env)
        self.assertNotIn("NCBI_API_KEY", env)

    def test_build_runpod_dockerstart_gzips_embeds_and_checks_size(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            template = tmp_path / "dockerstart.template.sh"
            pipeline = tmp_path / "run.sh"
            query = tmp_path / "query.faa"
            out = tmp_path / "dockerstart.sh"
            manifest = tmp_path / "manifest.json"
            template.write_text(
                "PIPE='__PIPELINE_B64__'\nQUERY='__QUERY_FASTA_B64__'\n",
                encoding="utf-8",
            )
            pipeline.write_text("#!/usr/bin/env bash\necho hello\n", encoding="utf-8")
            query.write_text(">q1\nMAAAAA\n", encoding="utf-8")
            result = build_runpod_dockerstart.build_dockerstart(
                template=template,
                pipeline=pipeline,
                inputs=[("QUERY_FASTA", query)],
                out=out,
                max_bytes=4096,
            )
            manifest.write_text(json.dumps(result), encoding="utf-8")
            rendered = out.read_text(encoding="utf-8")
            manifest_data = json.loads(manifest.read_text(encoding="utf-8"))
        self.assertTrue(result["ok"])
        self.assertNotIn("__PIPELINE_B64__", rendered)
        self.assertNotIn("__QUERY_FASTA_B64__", rendered)
        self.assertIn("output_bytes", manifest_data)

    def test_build_runpod_dockerstart_rejects_payload_over_size(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            template = tmp_path / "dockerstart.template.sh"
            pipeline = tmp_path / "run.sh"
            out = tmp_path / "dockerstart.sh"
            template.write_text("__PIPELINE_B64__\n", encoding="utf-8")
            pipeline.write_text("x" * 2000, encoding="utf-8")
            with self.assertRaises(ValueError):
                build_runpod_dockerstart.build_dockerstart(
                    template=template,
                    pipeline=pipeline,
                    out=out,
                    max_bytes=20,
                )

    def test_runpod_payload_preflight_rejects_first_boot_package_install(self) -> None:
        payload = {"dockerStartCmd": ["bash", "-lc", "mamba install -y hisat2 stringtie"], "env": {}}
        result = genecluster_runpod_rest_launch.payload_preflight(payload, warn_bytes=50 * 1024, max_bytes=60 * 1024)
        self.assertFalse(result["ok"])
        self.assertTrue(any("first-boot package installation" in error for error in result["errors"]))

    def test_orchestration_payload_preflight_rejects_first_boot_package_install(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            payload = Path(tmp) / "payload.json"
            payload.write_text(
                json.dumps({"dockerStartCmd": ["bash", "-lc", "set -e; apt-get install -y samtools"]}),
                encoding="utf-8",
            )
            result = symphony_orchestration_preflight.validate_payload(
                payload,
                warn_bytes=50 * 1024,
                max_bytes=60 * 1024,
            )
        self.assertFalse(result["ok"])
        self.assertTrue(any("first-boot package installation" in error for error in result["errors"]))

    def test_recovery_ledger_requires_degraded_reason(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            ledger = Path(tmp) / "recovery.json"
            ledger.write_text(
                json.dumps(
                    {
                        "issue_id": "GC-1",
                        "last_confirmed_state": "payload_preflight_failed",
                        "checkpoint_artifacts": ["prompt.md"],
                        "resume_command": "rerun preflight",
                        "degraded_recovery": True,
                    }
                ),
                encoding="utf-8",
            )
            result = symphony_orchestration_preflight.validate_recovery_ledger(ledger)
        self.assertFalse(result["ok"])
        self.assertTrue(any("degraded_reason" in error for error in result["errors"]))

    def test_sra_runinfo_parses_single_end_srx_to_srr(self) -> None:
        xml = """<?xml version="1.0"?>
<RunBundle>
  <EXPERIMENT accession="SRX1">
    <STUDY_REF accession="SRP1"/>
    <DESIGN><LIBRARY_DESCRIPTOR><LIBRARY_LAYOUT><SINGLE/></LIBRARY_LAYOUT></LIBRARY_DESCRIPTOR></DESIGN>
    <PLATFORM><ILLUMINA instrument_model="HiSeq 4000"/></PLATFORM>
  </EXPERIMENT>
  <SAMPLE accession="SRS1"><SAMPLE_NAME><TAXON_ID>123</TAXON_ID></SAMPLE_NAME></SAMPLE>
  <RUN_SET><RUN accession="SRR1" total_spots="10" total_bases="500" size="12345"/></RUN_SET>
</RunBundle>"""
        parsed = genecluster_sra_runinfo.parse_sra_xml(xml, input_accession="SRX1", dataset_id="rna1")
        row = parsed["runs"][0]
        self.assertEqual("SRR1", row["run_accession"])
        self.assertEqual("SINGLE", row["library_layout"])
        self.assertEqual("single_end", parsed["layout_branch"])
        self.assertEqual("single_end", row["layout_branch"])
        self.assertIn("hisat2 -U", row["aligner_hint"])

    def test_sra_runinfo_parses_paired_layout(self) -> None:
        xml = """<?xml version="1.0"?>
<RunBundle>
  <EXPERIMENT accession="SRX2">
    <DESIGN><LIBRARY_DESCRIPTOR><LIBRARY_LAYOUT><PAIRED/></LIBRARY_LAYOUT></LIBRARY_DESCRIPTOR></DESIGN>
    <PLATFORM><ILLUMINA instrument_model="NovaSeq 6000"/></PLATFORM>
  </EXPERIMENT>
  <RUN_SET><RUN accession="SRR2" total_spots="10" total_bases="3000" size="999"/></RUN_SET>
</RunBundle>"""
        parsed = genecluster_sra_runinfo.parse_sra_xml(xml, input_accession="SRX2")
        row = parsed["runs"][0]
        self.assertEqual("PAIRED", row["library_layout"])
        self.assertEqual("paired_end", parsed["layout_branch"])
        self.assertEqual("paired_end", row["layout_branch"])
        self.assertIn("_1.fastq", row["expected_fastq"])
        self.assertIn("_2.fastq", row["expected_fastq"])

    def test_sra_runinfo_writes_read_accessions_contract(self) -> None:
        xml = """<?xml version="1.0"?>
<RunBundle>
  <EXPERIMENT accession="SRX1">
    <STUDY_REF accession="SRP1"/>
    <DESIGN><LIBRARY_DESCRIPTOR><LIBRARY_LAYOUT><SINGLE/></LIBRARY_LAYOUT></LIBRARY_DESCRIPTOR></DESIGN>
    <PLATFORM><ILLUMINA instrument_model="HiSeq 4000"/></PLATFORM>
  </EXPERIMENT>
  <SAMPLE accession="SRS1"><SAMPLE_NAME><TAXON_ID>123</TAXON_ID></SAMPLE_NAME></SAMPLE>
  <RUN_SET><RUN accession="SRR1" total_spots="10" total_bases="500" size="12345"/></RUN_SET>
</RunBundle>"""
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            cache = tmp_path / "cache"
            cache.mkdir()
            (cache / "SRX1.xml").write_text(xml, encoding="utf-8")
            ledger = tmp_path / "data-ledger.tsv"
            ledger.write_text(
                "dataset_id\taccession\tdata_role\torganism\tremote_path\traw_artifact_policy\n"
                "reads\tSRX1\ttranscriptome_rna\tCoptis chinensis\t/workspace/genecluster/runs/test/inputs/sra/SRR1\tremote_only\n",
                encoding="utf-8",
            )
            out = tmp_path / "runinfo"
            rc = genecluster_sra_runinfo.main(
                [
                    "--data-ledger",
                    str(ledger),
                    "--xml-cache-dir",
                    str(cache),
                    "--out-dir",
                    str(out),
                    "--sleep-seconds",
                    "0",
                ]
            )
            rows, _fields = genecluster_preflight.read_tsv(out / "read-accessions.tsv")
            result = genecluster_preflight.validate_read_accessions(out / "read-accessions.tsv")

        self.assertEqual(0, rc)
        self.assertEqual([], result["errors"])
        self.assertEqual("reads:SRR1", rows[0]["source_id"])
        self.assertEqual("read_acquisition", rows[0]["source_record_type"])
        self.assertEqual("single_end", rows[0]["layout_branch"])
        self.assertEqual("metadata_resolved_raw_remote_only", rows[0]["acquisition_policy"])

    def test_sra_runinfo_keeps_bioproject_rows_visible_for_resolution(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            ledger = Path(tmp) / "data-ledger.tsv"
            ledger.write_text(
                "dataset_id\taccession\trole\n"
                "reads\tBIOPROJECT_UID_222281\tpublic transcriptome\n"
                "assembly\tGCA_000000000.1\tgenome assembly\n",
                encoding="utf-8",
            )
            rows = genecluster_sra_runinfo.rows_from_args(ledger, [])
        self.assertEqual(1, len(rows))
        self.assertEqual("BIOPROJECT_UID_222281", rows[0]["accession"])

    def test_stage_contract_can_validate_expected_output_artifacts(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            contract = tmp_path / "stage-contract.json"
            artifacts = tmp_path / "artifacts"
            (artifacts / "summary").mkdir(parents=True)
            (artifacts / "summary" / "candidate_hits.tsv").write_text("candidate_id\nc1\n", encoding="utf-8")
            contract.write_text(
                json.dumps(
                    {
                        "schema_version": 1,
                        "run_id": "r1",
                        "provider_class": "runpod_pod",
                        "run_scope": "candidate_search",
                        "progress_ledger": "stage-progress.jsonl",
                        "heartbeat_interval_minutes": 5,
                        "stale_after_minutes": 15,
                        "stages": [
                            {
                                "stage_id": "candidate-search",
                                "run_flag": "--candidate-search",
                                "entrypoint": "run.sh",
                                "expected_outputs": ["candidate_hits.tsv"],
                                "done_marker": "candidate.done",
                                "timeout_minutes": 30,
                                "resume_strategy": "idempotent rerun",
                                "failure_policy": "fail-closed",
                            }
                        ],
                        "watcher": {"required_for_runtime_hours_over": 1, "poll_interval_minutes": 5},
                        "acceptance": {
                            "partial_verdict_allowed": True,
                            "final_success_requires": ["candidate_hits.tsv"],
                            "partial_verdict_requires": ["run_summary.json"],
                        },
                    }
                ),
                encoding="utf-8",
            )
            result = genecluster_stage_contract.validate_expected_outputs(contract, artifacts)
        self.assertTrue(result["ok"])
        self.assertEqual(1, len(result["checked"]))

    def test_stage_contract_rejects_missing_expected_output_artifacts(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            contract = tmp_path / "stage-contract.json"
            contract.write_text(
                json.dumps(
                    {
                        "stages": [
                            {
                                "stage_id": "target-db-build",
                                "expected_outputs": ["target-db-indexes.tsv"],
                            }
                        ]
                    }
                ),
                encoding="utf-8",
            )
            result = genecluster_stage_contract.validate_expected_outputs(contract, tmp_path / "artifacts")
        self.assertFalse(result["ok"])
        self.assertTrue(any("expected output missing" in error for error in result["errors"]))

    def test_broad_workflow_class_bundle_emits_contracts(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            campaign_dir = tmp_path / "generic-campaign"
            campaign_dir.mkdir()
            for name in [
                "campaign-manifest.json",
                "data-ledger.tsv",
                "query-ledger.tsv",
                "resource-ledger.tsv",
                "project-goals.yaml",
                "pathway-steps.tsv",
                "database-ledger.tsv",
                "cache-ledger.tsv",
            ]:
                shutil.copyfile(EXAMPLE / name, campaign_dir / name)
            campaign_data = json.loads((campaign_dir / "campaign-manifest.json").read_text(encoding="utf-8"))
            campaign_data["campaign_id"] = "genecluster-generic-cross-species-v0"
            campaign_data["organism"] = "Target species B"
            campaign_data["target_pathway"] = "generic pathway A-to-Z"
            (campaign_dir / "campaign-manifest.json").write_text(json.dumps(campaign_data), encoding="utf-8")
            manifest = genecluster_launch_bundle.build_launch_bundle(
                campaign_path=campaign_dir / "campaign-manifest.json",
                out=tmp_path / "bundle",
                provider_class="local_full",
                run_scope="full_public_mining",
                repo_root=ROOT,
                heavy_workdir=str(tmp_path / "heavy" / "runs" / "generic"),
                run_id="generic-cross-species-workflow-classes",
                allow_local_full=True,
            )
            bundle = manifest.parent
            workflow_plan = json.loads((bundle / "workflow-class-plan.json").read_text(encoding="utf-8"))
            lane_plan = json.loads((bundle / "lane-activation-plan.json").read_text(encoding="utf-8"))
            prompt = (bundle / "campaign-prompt.md").read_text(encoding="utf-8")
            checks = [
                genecluster_preflight.validate_workflow_class_plan(bundle / "workflow-class-plan.json"),
                genecluster_preflight.validate_lane_activation_plan(bundle / "lane-activation-plan.json"),
                genecluster_preflight.validate_evidence_escalation_plan(bundle / "evidence-escalation-plan.json"),
                genecluster_preflight.validate_claim_levels(bundle / "claim-levels.tsv"),
                genecluster_preflight.validate_workflow_deferred_lanes(bundle / "workflow-deferred-lanes.tsv"),
                genecluster_preflight.validate_launch_manifest(manifest, repo_root=ROOT),
            ]
        errors = [error for result in checks for error in result["errors"]]
        self.assertEqual([], errors)
        self.assertEqual(genecluster_preflight.WORKFLOW_CLASSES, {row["workflow_class"] for row in workflow_plan["workflow_classes"]})
        self.assertIn("reference_first_anchor_mining", lane_plan["activated_lanes"])
        self.assertIn("singlecell_spatial_context", lane_plan["blocked_lanes"])
        self.assertIn("Workflow-class plan", prompt)

    def test_target_db_plan_rejects_repo_local_provider_paths(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            plan = {
                "schema_version": 1,
                "run_scope": "candidate_search",
                "data_ledger": "ledgers/data-ledger.tsv",
                "provider_db_root": str(ROOT / "bad-target-db"),
                "records": [
                    {
                        "target_db_id": "target_bad",
                        "dataset_id": "bad",
                        "species": "Species B",
                        "resource_kind": "target_protein_fasta",
                        "sequence_type": "protein",
                        "source_path": "/workspace/genecluster/runs/<run_id>/inputs/bad.faa",
                        "provider_path": str(ROOT / "bad-target-db" / "bad"),
                        "index_policy": "blast_diamond_mmseqs",
                        "local_copy": False,
                    }
                ],
                "index_targets": [],
                "outputs": {
                    "build_summary": "/workspace/genecluster/runs/test/summary/target-db-build-summary.json",
                    "resolved_ledger": "/workspace/genecluster/runs/test/summary/target-db-ledger.resolved.tsv",
                    "index_ledger": "/workspace/genecluster/runs/test/summary/target-db-indexes.tsv",
                },
                "local_copy": False,
            }
            path = Path(tmp) / "target-db-plan.json"
            path.write_text(json.dumps(plan), encoding="utf-8")
            result = genecluster_preflight.validate_target_db_plan(path, repo_root=ROOT)
        self.assertFalse(result["ok"])
        self.assertTrue(any("repo root" in error for error in result["errors"]))

    def test_database_scope_gates_keep_optional_max_out_of_candidate_plan(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            manifest = genecluster_launch_bundle.build_launch_bundle(
                campaign_path=EXAMPLE / "campaign-manifest.json",
                out=Path(tmp) / "bundle",
                provider_class="runpod_pod",
                run_scope="candidate_search",
                repo_root=ROOT,
                run_id="candidate-db-gates",
            )
            plan = json.loads((manifest.parent / "search-plan.json").read_text(encoding="utf-8"))
        self.assertIn("diamond_swissprot", plan["database_ids"])
        self.assertIn("hmmer_pfam", plan["database_ids"])
        self.assertNotIn("blast_nr", plan["database_ids"])
        self.assertIn("blast_nr", plan["optional_database_ids"])

    def test_one_day_full_bundle_has_runtime_budget_and_high_roi_gates(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            manifest = genecluster_launch_bundle.build_launch_bundle(
                campaign_path=EXAMPLE / "campaign-manifest.json",
                out=Path(tmp) / "bundle",
                provider_class="runpod_pod",
                run_scope="full_campaign_24h",
                repo_root=ROOT,
                run_id="one-day-full",
            )
            bundle = manifest.parent
            launch = json.loads(manifest.read_text(encoding="utf-8"))
            plan = json.loads((bundle / "search-plan.json").read_text(encoding="utf-8"))
            economics = json.loads((bundle / "run-economics.json").read_text(encoding="utf-8"))
            result = genecluster_preflight.validate_launch_manifest(manifest, repo_root=ROOT)
        self.assertEqual([], result["errors"])
        self.assertEqual("full_campaign_24h", launch["run_scope"])
        self.assertEqual(24, launch["runtime_policy"]["hard_stop_hours"])
        self.assertIn("--max-runtime-hours", launch["runner"]["command"])
        self.assertIn("diamond_refseq_plant", plan["database_ids"])
        self.assertIn("mibig_4_proteins", plan["database_ids"])
        self.assertIn("mmseqs_cdd", plan["database_ids"])
        self.assertNotIn("blast_refseq_protein", plan["database_ids"])
        self.assertNotIn("diamond_uniprot_trembl_plants", plan["database_ids"])
        self.assertNotIn("plantismash_2_resources", plan["database_ids"])
        self.assertEqual(24, economics["runtime_budget"]["target_runtime_hours"])
        self.assertIn("deferred-lanes.json", [item["path"] for item in launch["expected_artifacts"]])

    def test_candidate_route_plan_marks_transcript_first_gap(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            manifest = genecluster_launch_bundle.build_launch_bundle(
                campaign_path=EXAMPLE / "campaign-manifest.json",
                out=Path(tmp) / "bundle",
                provider_class="runpod_pod",
                run_scope="full_campaign_24h",
                repo_root=ROOT,
                run_id="route-gap",
            )
            bundle = manifest.parent
            plan = json.loads((bundle / "candidate-route-plan.json").read_text(encoding="utf-8"))
            preflight_result = genecluster_preflight.validate_candidate_route_plan(bundle / "candidate-route-plan.json")
            audit_result = genecluster_route_audit.audit_route_plan(plan)
            strict_result = genecluster_route_audit.audit_route_plan(plan, require_transcript_first=True)
        self.assertEqual([], preflight_result["errors"])
        self.assertEqual("transcript_first_then_genome_anchor", plan["primary_route"])
        self.assertEqual("rescue_only_not_primary_when_transcript_evidence_exists", plan["direct_genome_tblastn_policy"])
        self.assertIn("transcript_first_route_not_implemented_in_current_runner", plan["strict_scientific_blockers"])
        self.assertTrue(audit_result["ok"])
        self.assertFalse(strict_result["ok"])
        self.assertTrue(any("missing stages" in error for error in strict_result["errors"]))

    def test_db_bootstrap_plan_scope_gates_optional_max(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            manifest = genecluster_launch_bundle.build_launch_bundle(
                campaign_path=EXAMPLE / "campaign-manifest.json",
                out=Path(tmp) / "bundle",
                provider_class="runpod_pod",
                run_scope="candidate_search",
                repo_root=ROOT,
                run_id="candidate-db-bootstrap-gates",
            )
            plan = json.loads((manifest.parent / "db-bootstrap-plan.json").read_text(encoding="utf-8"))
            result = genecluster_preflight.validate_db_bootstrap_plan(manifest.parent / "db-bootstrap-plan.json")
        db_ids = {row["db_id"] for row in plan["records"]}
        self.assertIn("diamond_swissprot", db_ids)
        self.assertIn("hmmer_pfam", db_ids)
        self.assertNotIn("blast_nr", db_ids)
        self.assertIn("blast_nr", plan["optional_max_deferred"])
        self.assertEqual([], result["errors"])

    def test_decoy_plan_and_economics_surface_first_run_risks(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            manifest = genecluster_launch_bundle.build_launch_bundle(
                campaign_path=EXAMPLE / "campaign-manifest.json",
                out=Path(tmp) / "bundle",
                provider_class="runpod_pod",
                run_scope="candidate_search",
                repo_root=ROOT,
                run_id="candidate-risk-plans",
            )
            decoy_plan = json.loads((manifest.parent / "decoy-plan.json").read_text(encoding="utf-8"))
            economics = json.loads((manifest.parent / "run-economics.json").read_text(encoding="utf-8"))
        self.assertIn("Q003", decoy_plan["high_false_positive_risk_query_ids"])
        self.assertTrue(economics["cache_budget"]["search_result_cache_enabled"])
        self.assertIn("unresolved_high_confidence_query:Q002", economics["launch_blockers"])
        self.assertIn("diamond_swissprot", economics["database_budget"]["high_roi_required_database_ids"])

    def test_query_resolution_plan_flags_unresolved_public_seeds(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            manifest = genecluster_launch_bundle.build_launch_bundle(
                campaign_path=EXAMPLE / "campaign-manifest.json",
                out=Path(tmp) / "bundle",
                provider_class="runpod_pod",
                run_scope="candidate_search",
                repo_root=ROOT,
                run_id="query-resolution-plan",
            )
            plan_path = manifest.parent / "query-resolution-plan.json"
            plan = json.loads(plan_path.read_text(encoding="utf-8"))
            result = genecluster_preflight.validate_query_resolution_plan(plan_path, execution_ready=True)
        self.assertIn("Q002", plan["blocking_unresolved_query_ids"])
        self.assertIn("Q001", plan["warning_unresolved_query_ids"])
        self.assertFalse(result["ok"])
        self.assertTrue(any("unresolved high-confidence seeds" in error for error in result["errors"]))

    def test_remote_runner_cache_preflight_uses_scope_gates(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            manifest = genecluster_launch_bundle.build_launch_bundle(
                campaign_path=EXAMPLE / "campaign-manifest.json",
                out=tmp_path / "bundle",
                provider_class="local_full",
                run_scope="candidate_search",
                repo_root=ROOT,
                heavy_workdir=str(tmp_path / "heavy" / "runs" / "candidate"),
                run_id="candidate-cache-gates",
                allow_local_full=True,
            )
            data = json.loads(manifest.read_text(encoding="utf-8"))
            paths = genecluster_remote_runner.ensure_remote_layout(Path(data["heavy_workdir"]))
            result = genecluster_remote_runner.cache_preflight(data, manifest, paths, mock_tools=False)
        self.assertIn("diamond_swissprot", result["required_databases_missing"])
        self.assertNotIn("blast_refseq_protein", result["required_databases_missing"])
        self.assertNotIn("plantismash_2_resources", result["required_databases_missing"])

    def test_strict_resource_review_promotes_warning_to_error(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "resource-ledger.tsv"
            path.write_text(
                "resource\tresource_type\tversion\tlicense_class\tuse_mode\tcitation\n"
                "RestrictedDB\tpathway_database\taccess_date\trestricted-or-review\treviewed_reference\thttps://example.org\n",
                encoding="utf-8",
            )
            relaxed = genecluster_preflight.validate_resource_ledger(path, strict=False)
            strict = genecluster_preflight.validate_resource_ledger(path, strict=True)
        self.assertEqual([], relaxed["errors"])
        self.assertTrue(relaxed["warnings"])
        self.assertFalse(strict["ok"])
        self.assertTrue(any("requires explicit approval" in error for error in strict["errors"]))

    def test_local_artifact_scan_blocks_raw_sequence_extensions(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            (root / "candidate.fasta").write_text(">x\nATGC\n", encoding="utf-8")
            result = genecluster_preflight.scan_local_artifacts(root)
        self.assertFalse(result["ok"])
        self.assertTrue(any("candidate.fasta" in error for error in result["errors"]))

    def test_launch_manifest_hash_mismatch_fails(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            manifest = genecluster_launch_bundle.build_launch_bundle(
                campaign_path=EXAMPLE / "campaign-manifest.json",
                out=Path(tmp) / "bundle",
                provider_class="runpod_pod",
                run_scope="smoke",
                repo_root=ROOT,
                run_id="hash-check",
            )
            data = json.loads(manifest.read_text(encoding="utf-8"))
            data["artifact_policy"] = "tampered"
            manifest.write_text(json.dumps(data), encoding="utf-8")
            result = genecluster_preflight.validate_launch_manifest(manifest, repo_root=ROOT)
        self.assertFalse(result["ok"])
        self.assertTrue(any("launch_payload_sha256 mismatch" in error for error in result["errors"]))

    def test_full_issue_dry_run_outputs_valid_contracts(self) -> None:
        issues = genecluster_issue_dry_run.build_issues(
            EXAMPLE / "campaign-manifest.json",
            "GENECLUSTER-MIT",
            "full_campaign",
        )
        self.assertIn("GENECLUSTER-MIT-W11-FULL-CAMPAIGN-AUDIT", issues)
        for body in issues.values():
            result = genecluster_issue_dry_run.validate_issue(body)
            self.assertEqual([], result["errors"])

    def test_one_day_full_issue_dry_run_outputs_valid_contracts(self) -> None:
        issues = genecluster_issue_dry_run.build_issues(
            EXAMPLE / "campaign-manifest.json",
            "GENECLUSTER-MIT",
            "full_campaign_24h",
        )
        self.assertIn("GENECLUSTER-MIT-W05-FULL-CAMPAIGN-24H-BUNDLE", issues)
        self.assertIn("GENECLUSTER-MIT-W11-FULL-CAMPAIGN-24H-AUDIT", issues)
        for body in issues.values():
            result = genecluster_issue_dry_run.validate_issue(body)
            self.assertEqual([], result["errors"])

    def test_claim_auditor_catches_core_overclaims(self) -> None:
        rows = [
            {
                "candidate_id": "BAD_TRANSCRIPT_CLUSTER",
                "hit_type": "transcript_hit",
                "product_claim_level": "cluster_hypothesis",
                "domain_calls": "MDR",
                "domain_architecture": "MDR",
                "duplicate_class": "representative",
                "evidence_score": "0.8",
                "pathway_step_id": "STEP_DCS",
                "genome_locus": "transcript_only",
                "neighborhood_cluster_id": "remote_pending",
                "review_status": "needs-human-review",
            },
            {
                "candidate_id": "BAD_BROAD_CYP",
                "hit_type": "domain_hit",
                "product_claim_level": "pathway_hypothesis",
                "domain_calls": "p450",
                "domain_architecture": "CYP450",
                "duplicate_class": "broad_family",
                "evidence_score": "0.4",
                "pathway_step_id": "STEP_CYP_HYDROXYLATION",
                "genome_locus": "remote_pending",
                "neighborhood_cluster_id": "remote_pending",
                "review_status": "needs-human-review",
            },
        ]
        records = genecluster_claim_audit.audit_candidates(rows, campaign_id="test")
        rule_ids = {record["rule_id"] for record in records}
        self.assertIn("transcriptome_only_does_not_prove_physical_cluster", rule_ids)
        self.assertIn("broad_family_hit_does_not_prove_product_chemistry", rule_ids)
        self.assertIn("product_claim_requires_external_validation", rule_ids)

    def test_remote_runner_mock_outputs_validate(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            manifest = genecluster_launch_bundle.build_launch_bundle(
                campaign_path=EXAMPLE / "campaign-manifest.json",
                out=tmp_path / "bundle",
                provider_class="local_full",
                run_scope="candidate_search",
                repo_root=ROOT,
                heavy_workdir=str(tmp_path / "heavy" / "runs" / "candidate"),
                run_id="remote-runner-mock",
                allow_local_full=True,
            )
            out = tmp_path / "summary"
            genecluster_remote_runner.run(
                manifest,
                out,
                toolcheck=True,
                cache_preflight_flag=True,
                query_preflight_flag=True,
                decoy_preflight_flag=True,
                candidate_search=True,
                mock_tools=True,
            )
            checks = [
                genecluster_preflight.validate_candidate_hits(out / "candidate_hits.tsv"),
                genecluster_preflight.validate_jsonl(
                    out / "evidence.jsonl",
                    required_keys={"claim_id", "subject_id", "evidence_class", "source_artifact", "confidence", "review_status"},
                    label="evidence.jsonl",
                ),
                genecluster_preflight.validate_claim_audit_jsonl(out / "claim-audit.jsonl"),
                genecluster_preflight.validate_jsonl(
                    out / "provenance.jsonl",
                    required_keys={"kind", "campaign_id"},
                    label="provenance.jsonl",
                ),
            ]
            self.assertTrue((out / "query-preflight.json").exists())
            self.assertTrue((out / "decoy-preflight.json").exists())
            self.assertTrue((out / "evidence.sqlite").exists())
            self.assertTrue((out / "search-cache-manifest.json").exists())
            self.assertTrue((out / "stage-progress.jsonl").exists())
            progress_result = genecluster_stage_contract.validate_progress_ledger(out / "stage-progress.jsonl", require_terminal=True)
            query_summary = json.loads((out / "query-preflight.json").read_text(encoding="utf-8"))
            self.assertTrue(query_summary["query_fasta_present"])
            import sqlite3

            with closing(sqlite3.connect(out / "evidence.sqlite")) as conn:
                count = conn.execute("SELECT COUNT(*) FROM candidate_hits").fetchone()[0]
            self.assertEqual(1, count)
        errors = [error for result in checks for error in result["errors"]]
        errors.extend(progress_result["errors"])
        self.assertEqual([], errors)

    def test_contract_self_check_rejects_mock_target_success(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            manifest = genecluster_launch_bundle.build_launch_bundle(
                campaign_path=EXAMPLE / "campaign-manifest.json",
                out=tmp_path / "bundle",
                provider_class="local_full",
                run_scope="full_campaign_24h",
                repo_root=ROOT,
                heavy_workdir=str(tmp_path / "heavy" / "runs" / "selfcheck-mock"),
                run_id="selfcheck-mock",
                allow_local_full=True,
                allow_provider_large_downloads=True,
            )
            out = tmp_path / "summary"
            genecluster_remote_runner.run(
                manifest,
                out,
                toolcheck=True,
                db_bootstrap=True,
                data_materialization=True,
                target_db_build=True,
                query_preflight_flag=True,
                candidate_search=True,
                full_campaign=True,
                mock_tools=True,
                max_runtime_hours=24,
            )
            proc = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPTS / "genecluster_contract_self_check.py"),
                    "--summary-dir",
                    str(out),
                    "--require-real-target-search",
                    "--json",
                ],
                check=False,
                capture_output=True,
                text=True,
            )
            result = json.loads(proc.stdout)
        self.assertNotEqual(0, proc.returncode)
        self.assertFalse(result["ok"])
        self.assertTrue(any("mock" in error for error in result["errors"]))
        self.assertTrue(any("target_* databases" in error for error in result["errors"]))

    def test_contract_self_check_accepts_joined_real_target_summary(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp)
            genecluster_remote_runner.write_json(
                out / "run_summary.json",
                {
                    "toolcheck_ok": True,
                    "data_materialization_ok": True,
                    "target_db_build_ok": True,
                    "candidate_search_ok": True,
                    "real_target_search_ok": True,
                    "heavy_execution_performed": True,
                },
            )
            genecluster_remote_runner.write_json(out / "data-materialization-summary.json", {"ok": True, "dry_run": False, "mock_tools": False})
            genecluster_remote_runner.write_json(out / "target-db-build-summary.json", {"ok": True, "dry_run": False, "mock_tools": False})
            genecluster_remote_runner.write_json(
                out / "candidate-search-summary.json",
                {"ok": True, "mock_tools": False, "real_target_search_ok": True, "target_commands_completed": 1},
            )
            genecluster_remote_runner.write_tsv(
                out / "materialized-targets.tsv",
                [
                    "dataset_id",
                    "accession",
                    "run_id",
                    "data_role",
                    "organism",
                    "materialization_status",
                    "source_kind",
                    "target_fasta",
                    "fastq_dir",
                    "read_count",
                    "base_count",
                    "command_summary",
                    "review_status",
                ],
                [
                    {
                        "dataset_id": "target_reads",
                        "accession": "SRX1",
                        "materialization_status": "target_fasta_materialized",
                        "target_fasta": "/workspace/genecluster/runs/test/inputs/target-sequences/target_reads/target_sequences.fasta",
                    }
                ],
            )
            genecluster_remote_runner.write_tsv(
                out / "target-db-indexes.tsv",
                [
                    "target_db_id",
                    "dataset_id",
                    "engine",
                    "sequence_type",
                    "index_path",
                    "source_path",
                    "build_status",
                    "command",
                    "checksum_status",
                ],
                [
                    {
                        "target_db_id": "target_target_reads",
                        "dataset_id": "target_reads",
                        "engine": "blast",
                        "sequence_type": "nucleotide",
                        "index_path": "/workspace/genecluster/runs/test/databases/target/target_reads/indexes/blast",
                        "source_path": "/workspace/genecluster/runs/test/inputs/target-sequences/target_reads/target_sequences.fasta",
                        "build_status": "built",
                    }
                ],
            )
            row = dict(genecluster_remote_runner.mock_candidate_rows()[0])
            row.update(
                {
                    "candidate_id": "GCAND_REAL_0001",
                    "gene_or_transcript_id": "target_read_001",
                    "dataset_id": "target_reads",
                    "target_db_id": "target_target_reads",
                    "target_species": "Target species",
                }
            )
            genecluster_remote_runner.write_tsv(out / "candidate_hits.tsv", genecluster_remote_runner.CANDIDATE_HEADERS, [row])
            proc = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPTS / "genecluster_contract_self_check.py"),
                    "--summary-dir",
                    str(out),
                    "--require-real-target-search",
                    "--json",
                ],
                check=False,
                capture_output=True,
                text=True,
            )
            result = json.loads(proc.stdout)
        self.assertEqual(0, proc.returncode)
        self.assertTrue(result["ok"], result)
        self.assertEqual(1, result["target_candidate_rows"])

    def test_sra_materialization_uses_downloaded_sra_artifact_not_srx_path(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            heavy = tmp_path / "heavy" / "runs" / "sra-test"
            calls: list[list[str]] = []

            def fake_run_command(command: list[str], *, timeout_seconds: int) -> dict[str, object]:
                calls.append(command)
                if command[0] == "prefetch":
                    (heavy / "inputs" / "sra" / "SRX999" / "SRR123456").mkdir(parents=True, exist_ok=True)
                    (heavy / "inputs" / "sra" / "SRX999" / "SRR123456" / "SRR123456.sra").write_text("stub", encoding="utf-8")
                if command[0] == "fasterq-dump":
                    out_dir = Path(command[command.index("-O") + 1])
                    out_dir.mkdir(parents=True, exist_ok=True)
                    (out_dir / "SRR123456.fastq").write_text("@r1\nATGC\n+\nIIII\n", encoding="utf-8")
                return {"command": command, "returncode": 0}

            row = {
                "dataset_id": "target_reads",
                "accession": "SRX999",
                "run_id": "",
                "data_role": "RNA-seq transcriptome",
                "organism": "Target species",
            }
            with patch.object(genecluster_data_materialization.shutil, "which", return_value="/usr/bin/tool"):
                with patch.object(genecluster_data_materialization, "run_command", side_effect=fake_run_command):
                    materialized, _commands = genecluster_data_materialization.materialize_sra_row(
                        row,
                        run_id="sra-test",
                        heavy_workdir=heavy,
                        allow_large_downloads=True,
                        dry_run=False,
                        mock_tools=False,
                        timeout_seconds=60,
                    )

        fasterq_calls = [call for call in calls if call[0] == "fasterq-dump"]
        self.assertEqual("target_fasta_materialized", materialized["materialization_status"])
        self.assertEqual(1, len(fasterq_calls))
        self.assertTrue(fasterq_calls[0][1].endswith("SRR123456.sra"), fasterq_calls[0])
        self.assertNotEqual(str(heavy / "inputs" / "sra" / "SRX999"), fasterq_calls[0][1])

    def test_fastq_to_fasta_accepts_single_end_without_mate_file(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            fastq = tmp_path / "SRR123456_1.fastq"
            fastq.write_text("@r1\nATGC\n+\nIIII\n", encoding="utf-8")
            fasta = tmp_path / "target_sequences.fasta"
            read_count, base_count = genecluster_data_materialization.fastq_to_fasta([fastq], fasta)
            fasta_text = fasta.read_text(encoding="utf-8")

        self.assertEqual(1, read_count)
        self.assertEqual(4, base_count)
        self.assertIn(">SRR123456_1|1|r1", fasta_text)

    def test_remote_runner_full_mock_outputs_anchor_and_neighborhood_tables(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            manifest = genecluster_launch_bundle.build_launch_bundle(
                campaign_path=EXAMPLE / "campaign-manifest.json",
                out=tmp_path / "bundle",
                provider_class="local_full",
                run_scope="full_campaign",
                repo_root=ROOT,
                heavy_workdir=str(tmp_path / "heavy" / "runs" / "full"),
                run_id="remote-runner-full-mock",
                allow_local_full=True,
            )
            out = tmp_path / "summary"
            genecluster_remote_runner.run(
                manifest,
                out,
                toolcheck=True,
                db_bootstrap=True,
                cache_preflight_flag=True,
                reference_import=True,
                query_preflight_flag=True,
                decoy_preflight_flag=True,
                candidate_search=True,
                anchor_map=True,
                neighborhood_extract=True,
                mock_tools=True,
            )
            checks = [
                genecluster_preflight.validate_candidate_hits(out / "candidate_hits.tsv"),
                genecluster_preflight.validate_candidate_anchors(out / "candidate_anchors.tsv"),
                genecluster_preflight.validate_cluster_neighborhoods(out / "cluster_neighborhoods.tsv"),
                genecluster_preflight.validate_neighbor_annotations(out / "neighbor_annotations.tsv"),
                genecluster_preflight.validate_domain_labels(out / "domain_labels.tsv"),
            ]
            summary = json.loads((out / "run_summary.json").read_text(encoding="utf-8"))
            manifest_data = json.loads(manifest.read_text(encoding="utf-8"))
            missing_expected = [
                item["path"]
                for item in manifest_data["expected_artifacts"]
                if not (out / item["path"]).exists()
            ]
            with closing(sqlite3.connect(out / "evidence.sqlite")) as conn:
                anchor_count = conn.execute("SELECT COUNT(*) FROM candidate_anchors").fetchone()[0]
                neighborhood_count = conn.execute("SELECT COUNT(*) FROM cluster_neighborhoods").fetchone()[0]
        errors = [error for result in checks for error in result["errors"]]
        self.assertEqual([], errors)
        self.assertTrue(summary["db_bootstrap_ok"])
        self.assertTrue(summary["reference_import_ok"])
        self.assertTrue(summary["anchor_map_ok"])
        self.assertTrue(summary["neighborhood_extract_ok"])
        self.assertEqual([], missing_expected)
        self.assertEqual(1, anchor_count)
        self.assertEqual(1, neighborhood_count)

    def test_remote_runner_full_mock_outputs_superpower_tables(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            manifest = genecluster_launch_bundle.build_launch_bundle(
                campaign_path=EXAMPLE / "campaign-manifest.json",
                out=tmp_path / "bundle",
                provider_class="local_full",
                run_scope="full_campaign",
                repo_root=ROOT,
                heavy_workdir=str(tmp_path / "heavy" / "runs" / "full"),
                run_id="remote-runner-superpower-mock",
                allow_local_full=True,
            )
            out = tmp_path / "summary"
            genecluster_remote_runner.run(
                manifest,
                out,
                toolcheck=True,
                db_bootstrap=True,
                target_db_build=True,
                cache_preflight_flag=True,
                reference_import=True,
                query_preflight_flag=True,
                decoy_preflight_flag=True,
                candidate_search=True,
                anchor_map=True,
                orthology_anchor=True,
                neighborhood_extract=True,
                neighborhood_score=True,
                pathway_completeness=True,
                mock_tools=True,
            )
            checks = [
                genecluster_preflight.validate_target_db_resolved(out / "target-db-ledger.resolved.tsv", repo_root=ROOT),
                genecluster_preflight.validate_target_db_indexes(out / "target-db-indexes.tsv", repo_root=ROOT),
                genecluster_preflight.validate_orthology_links(out / "orthology_links.tsv"),
                genecluster_preflight.validate_anchor_ladder(out / "anchor_ladder.tsv"),
                genecluster_preflight.validate_reciprocal_hits(out / "reciprocal_hits.tsv"),
                genecluster_preflight.validate_neighborhood_hypotheses(out / "neighborhood_hypotheses.tsv"),
                genecluster_preflight.validate_pathway_completeness(out / "pathway_completeness.tsv"),
            ]
            summary = json.loads((out / "run_summary.json").read_text(encoding="utf-8"))
            with closing(sqlite3.connect(out / "evidence.sqlite")) as conn:
                orthology_count = conn.execute("SELECT COUNT(*) FROM orthology_links").fetchone()[0]
                pathway_count = conn.execute("SELECT COUNT(*) FROM pathway_completeness").fetchone()[0]
        errors = [error for result in checks for error in result["errors"]]
        self.assertEqual([], errors)
        self.assertTrue(summary["target_db_build_ok"])
        self.assertTrue(summary["orthology_anchor_ok"])
        self.assertTrue(summary["neighborhood_score_ok"])
        self.assertTrue(summary["pathway_completeness_ok"])
        self.assertEqual(1, orthology_count)
        self.assertGreaterEqual(pathway_count, 1)

    def test_remote_runner_full_mock_outputs_workflow_class_tables(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            manifest = genecluster_launch_bundle.build_launch_bundle(
                campaign_path=EXAMPLE / "campaign-manifest.json",
                out=tmp_path / "bundle",
                provider_class="local_full",
                run_scope="full_public_mining",
                repo_root=ROOT,
                heavy_workdir=str(tmp_path / "heavy" / "runs" / "full-public"),
                run_id="remote-runner-workflow-class-mock",
                allow_local_full=True,
            )
            out = tmp_path / "summary"
            genecluster_remote_runner.run(
                manifest,
                out,
                toolcheck=True,
                db_bootstrap=True,
                target_db_build=True,
                cache_preflight_flag=True,
                reference_import=True,
                query_preflight_flag=True,
                decoy_preflight_flag=True,
                candidate_search=True,
                anchor_map=True,
                orthology_anchor=True,
                neighborhood_extract=True,
                neighborhood_score=True,
                pathway_completeness=True,
                workflow_classes=True,
                mock_tools=True,
            )
            checks = [
                genecluster_preflight.validate_claim_levels(out / "claim-levels.tsv"),
                genecluster_preflight.validate_workflow_deferred_lanes(out / "workflow-deferred-lanes.tsv"),
                genecluster_preflight.validate_summary_table(out / "isoform-ledger.tsv", required_columns=genecluster_preflight.ISOFORM_LEDGER_COLUMNS, label="isoform ledger", repo_root=ROOT),
                genecluster_preflight.validate_summary_table(out / "transcriptome-build-ledger.tsv", required_columns=genecluster_preflight.TRANSCRIPTOME_BUILD_LEDGER_COLUMNS, label="transcriptome build ledger", repo_root=ROOT),
                genecluster_preflight.validate_summary_table(out / "copy-classification.tsv", required_columns=genecluster_preflight.COPY_CLASSIFICATION_COLUMNS, label="copy classification", repo_root=ROOT),
                genecluster_preflight.validate_summary_table(out / "expression-design.tsv", required_columns=genecluster_preflight.EXPRESSION_DESIGN_COLUMNS, label="expression design", repo_root=ROOT),
                genecluster_preflight.validate_expression_matrix_manifest(out / "expression-matrix-manifest.json", repo_root=ROOT),
                genecluster_preflight.validate_summary_table(out / "comparative_neighborhoods.tsv", required_columns=genecluster_preflight.COMPARATIVE_NEIGHBORHOOD_COLUMNS, label="comparative neighborhoods", repo_root=ROOT),
                genecluster_preflight.validate_summary_table(out / "pav-copy-number.tsv", required_columns=genecluster_preflight.PAV_COPY_NUMBER_COLUMNS, label="PAV copy number", repo_root=ROOT),
                genecluster_preflight.validate_summary_table(out / "candidate_interval_sv.tsv", required_columns=genecluster_preflight.CANDIDATE_INTERVAL_SV_COLUMNS, label="candidate interval SV", repo_root=ROOT),
                genecluster_preflight.validate_summary_table(out / "graph_path_support.tsv", required_columns=genecluster_preflight.GRAPH_PATH_SUPPORT_COLUMNS, label="graph path support", repo_root=ROOT),
                genecluster_preflight.validate_summary_table(out / "spatial-domain-expression.tsv", required_columns=genecluster_preflight.SPATIAL_DOMAIN_EXPRESSION_COLUMNS, label="spatial/domain expression", repo_root=ROOT),
                genecluster_preflight.validate_longread_qc(out / "longread-qc.json"),
            ]
            summary = json.loads((out / "run_summary.json").read_text(encoding="utf-8"))
            manifest_data = json.loads(manifest.read_text(encoding="utf-8"))
            missing_expected = [
                item["path"]
                for item in manifest_data["expected_artifacts"]
                if not (out / item["path"]).exists()
            ]
            with closing(sqlite3.connect(out / "evidence.sqlite")) as conn:
                isoform_count = conn.execute("SELECT COUNT(*) FROM isoform_ledger").fetchone()[0]
                graph_count = conn.execute("SELECT COUNT(*) FROM graph_path_support").fetchone()[0]
        errors = [error for result in checks for error in result["errors"]]
        self.assertEqual([], errors)
        self.assertTrue(summary["workflow_classes_ok"])
        self.assertEqual([], missing_expected)
        self.assertEqual(1, isoform_count)
        self.assertEqual(1, graph_count)

    def test_remote_runner_one_day_mock_outputs_deferred_lane_manifest(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            manifest = genecluster_launch_bundle.build_launch_bundle(
                campaign_path=EXAMPLE / "campaign-manifest.json",
                out=tmp_path / "bundle",
                provider_class="local_full",
                run_scope="full_campaign_24h",
                repo_root=ROOT,
                heavy_workdir=str(tmp_path / "heavy" / "runs" / "one-day"),
                run_id="remote-runner-one-day-mock",
                allow_local_full=True,
            )
            out = tmp_path / "summary"
            genecluster_remote_runner.run(
                manifest,
                out,
                toolcheck=True,
                db_bootstrap=True,
                cache_preflight_flag=True,
                reference_import=True,
                query_preflight_flag=True,
                decoy_preflight_flag=True,
                candidate_search=True,
                anchor_map=True,
                neighborhood_extract=True,
                full_campaign=True,
                mock_tools=True,
                max_runtime_hours=24,
            )
            summary = json.loads((out / "run_summary.json").read_text(encoding="utf-8"))
            deferred = json.loads((out / "deferred-lanes.json").read_text(encoding="utf-8"))
            completeness_result = genecluster_preflight.validate_pathway_completeness(
                out / "pathway_completeness.tsv",
                require_deferred_budget=True,
            )
            manifest_data = json.loads(manifest.read_text(encoding="utf-8"))
            missing_expected = [
                item["path"]
                for item in manifest_data["expected_artifacts"]
                if not (out / item["path"]).exists()
            ]
        self.assertEqual([], missing_expected)
        self.assertTrue(summary["runtime_budget"]["enabled"])
        self.assertEqual(24, summary["runtime_budget"]["max_runtime_hours"])
        self.assertIn("de_novo_transcriptome_assembly", deferred["deferred_by_policy"])
        self.assertEqual([], completeness_result["errors"])

    def test_workflow_deferred_lanes_require_budget_rows_for_one_day(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            manifest = genecluster_launch_bundle.build_launch_bundle(
                campaign_path=EXAMPLE / "campaign-manifest.json",
                out=Path(tmp) / "bundle",
                provider_class="runpod_pod",
                run_scope="full_campaign_24h",
                repo_root=ROOT,
                run_id="one-day-workflow-deferred",
            )
            result = genecluster_preflight.validate_workflow_deferred_lanes(
                manifest.parent / "workflow-deferred-lanes.tsv",
                require_deferred_budget=True,
            )
        self.assertEqual([], result["errors"])

    def test_workflow_summary_validator_rejects_repo_local_heavy_paths(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "assembly-ledger.tsv"
            genecluster_remote_runner.write_tsv(
                path,
                genecluster_remote_runner.ASSEMBLY_LEDGER_HEADERS,
                [
                    {
                        "assembly_id": "bad",
                        "species": "Target species",
                        "accession": "GCA_BAD",
                        "assembly_role": "target",
                        "remote_path": str(ROOT / "bad-local-genome.fasta"),
                        "coordinate_system": "bad",
                        "checksum_status": "remote_pending",
                        "review_status": "needs-human-review",
                    }
                ],
            )
            result = genecluster_preflight.validate_summary_table(
                path,
                required_columns=genecluster_preflight.ASSEMBLY_LEDGER_COLUMNS,
                label="assembly ledger",
                repo_root=ROOT,
            )
        self.assertFalse(result["ok"])
        self.assertTrue(any("must not be under the repo root" in error for error in result["errors"]))

    def test_candidate_preflight_enforces_core_claim_boundaries(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "candidate_hits.tsv"
            headers = list(genecluster_remote_runner.CANDIDATE_HEADERS)
            base = genecluster_remote_runner.mock_candidate_rows()[0]
            bad_transcript = dict(base, candidate_id="BAD_TRANSCRIPT", product_claim_level="cluster_hypothesis")
            bad_domain = dict(base, candidate_id="BAD_DOMAIN", hit_type="domain_hit", duplicate_class="broad_family", product_claim_level="pathway_hypothesis")
            genecluster_remote_runner.write_tsv(path, headers, [bad_transcript, bad_domain])
            result = genecluster_preflight.validate_candidate_hits(path)
        self.assertFalse(result["ok"])
        self.assertTrue(any("transcript_hit cannot carry cluster_hypothesis" in error for error in result["errors"]))
        self.assertTrue(any("broad-family/domain-only evidence" in error for error in result["errors"]))

    def test_reference_import_blocks_heavy_reference_downloads_without_opt_in(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            data = tmp_path / "data-ledger.tsv"
            data.write_text(
                "dataset_id\taccession\trun_id\tdata_role\tsample_type\torganism\tbioproject\ttechnology\texpected_size\tsource_url\tremote_path\tchecksum_status\tdata_sensitivity\tallowed_upload\traw_artifact_policy\toperator_approval_id\tlicense_url\tcitation_doi\tmd5_or_sha256\n"
                "asm\tGCA_000001\t\tassembly\tleaf\tTest organism\tPRJ\tassembly\t1 Gb\thttps://example.org\t/workspace/genecluster/runs/<run_id>/inputs/asm\tremote_pending\tpublic\tno_public_webserver_upload\tremote_only\tnot_required_public_data\thttps://example.org\t\t\n",
                encoding="utf-8",
            )
            resources = tmp_path / "resource-ledger.tsv"
            resources.write_text(
                "resource\tresource_type\tversion\tlicense_class\tuse_mode\tcitation\tapproval_status\tnotes\n"
                "NCBI Datasets CLI\treference_import\tremote-image-pinned\topen-data-with-terms\tremote_container\thttps://www.ncbi.nlm.nih.gov/datasets/docs/v2/command-line-tools/download-and-install/\tapproved\tfixture\n",
                encoding="utf-8",
            )
            manifest = tmp_path / "launch-manifest.json"
            manifest.write_text(
                json.dumps(
                    {
                        "run_id": "reference-gate",
                        "run_scope": "full_campaign",
                        "heavy_workdir": str(tmp_path / "heavy" / "runs" / "reference-gate"),
                        "data_ledger": str(data),
                        "resource_ledger": str(resources),
                    }
                ),
                encoding="utf-8",
            )
            summary_path = genecluster_reference_import.run(manifest, tmp_path / "summary", dry_run=True, mock_tools=False)
            summary = json.loads(summary_path.read_text(encoding="utf-8"))
        self.assertFalse(summary["ok"])
        self.assertEqual("blocked_reference_download_requires_opt_in", summary["records"][0]["status"])

    def test_provider_query_resolver_uses_public_protein_accessions(self) -> None:
        plan = {
            "records": [
                {
                    "query_id": "Q004",
                    "query_name": "CjBBE",
                    "resolution_action": "fetch_public_accession",
                    "resolved_accession": "mRNA:OQ129427.1;protein:WKU61907.1",
                }
            ]
        }

        def fake_fetcher(url: str) -> str:
            self.assertIn("WKU61907.1", url)
            self.assertNotIn("OQ129427.1", url)
            return ">WKU61907.1 mocked protein\nMSTNPKPQR\n"

        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "protein_queries.faa"
            result = genecluster_remote_runner.resolve_query_fasta_from_plan(plan, out, fetcher=fake_fetcher)
            text = out.read_text(encoding="utf-8")
        self.assertEqual(1, result["resolved_count"])
        self.assertIn(">Q004|CjBBE|WKU61907.1", text)

    def test_remote_runner_query_preflight_blocks_without_fasta(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            manifest = genecluster_launch_bundle.build_launch_bundle(
                campaign_path=EXAMPLE / "campaign-manifest.json",
                out=tmp_path / "bundle",
                provider_class="local_full",
                run_scope="candidate_search",
                repo_root=ROOT,
                heavy_workdir=str(tmp_path / "heavy" / "runs" / "candidate"),
                run_id="query-preflight-blocks",
                allow_local_full=True,
            )
            data = json.loads(manifest.read_text(encoding="utf-8"))
            paths = genecluster_remote_runner.ensure_remote_layout(Path(data["heavy_workdir"]))
            result = genecluster_remote_runner.query_preflight(data, manifest, paths, mock_tools=False)
        self.assertFalse(result["ok"])
        self.assertFalse(result["query_fasta_present"])
        self.assertIn("Q002", result["blocking_unresolved_query_ids"])


class GeneClusterAtlasSuperpowerTests(unittest.TestCase):
    def write_query_fasta(self, path: Path, include_controls: bool = True) -> None:
        if include_controls:
            text = (
                ">ACT2_control Arabidopsis actin 2 positive control\nMAAAAA\n"
                ">GAPDH_control glycolysis positive control\nMGGGGG\n"
                ">random_shuffle_negative shuffled decoy control\nMSSSSS\n"
                ">pathway_query strictosidine synthase\nMQQQQQ\n"
            )
        else:
            text = ">pathway_query strictosidine synthase\nMQQQQQ\n"
        path.write_text(text, encoding="utf-8")

    def write_annotation_pair(self, tmp_path: Path) -> tuple[Path, Path]:
        proteome = tmp_path / "proteome.faa"
        proteome.write_text(">protA annotated protein\nMAAAAA\n>protB neighbor\nMBBBBB\n", encoding="utf-8")
        gff = tmp_path / "genomic.gff"
        gff.write_text(
            "chr1\tRefSeq\tCDS\t10\t99\t.\t+\t0\tID=cds-protA;protein_id=protA;product=anchor\n"
            "chr1\tRefSeq\tCDS\t150\t240\t.\t+\t0\tID=cds-protB;protein_id=protB;product=neighbor\n",
            encoding="utf-8",
        )
        return proteome, gff

    def write_tsv(self, path: Path, headers: list[str], rows: list[dict[str, str]]) -> None:
        with path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, delimiter="\t", fieldnames=headers)
            writer.writeheader()
            writer.writerows(rows)

    def test_annotation_scout_selects_annotation_direct_for_joined_proteome_gff(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            query = tmp_path / "queries.faa"
            self.write_query_fasta(query)
            proteome, gff = self.write_annotation_pair(tmp_path)
            ledger = tmp_path / "source-ledger.tsv"
            ledger.write_text(
                "source_id\torganism\tproteome_fasta\tgff\tgenome_fasta\ttranscriptome\ttranscriptome_species\n"
                f"coptis\tCoptis chinensis\t{proteome}\t{gff}\t\t\tCoptis chinensis\n",
                encoding="utf-8",
            )
            sources = genecluster_annotation_scout.load_source_ledger(ledger)
            decision, rows = genecluster_annotation_scout.build_route_decision(query, sources)
            out_paths = genecluster_annotation_scout.write_route_outputs(decision, rows, tmp_path / "route-scout")
            wrote_decision = Path(out_paths["route_decision"]).exists()
            wrote_ledger = Path(out_paths["annotation_ledger"]).exists()

        self.assertEqual("annotation_direct", decision["selected_route"])
        self.assertEqual("L3_annotation_neighborhood_ready", decision["claim_ceiling"])
        self.assertEqual(2, rows[0]["protein_gff_join_count"])
        self.assertIn("ACT2", decision["positive_controls"])
        self.assertIn("random_shuffle", decision["negative_controls"])
        self.assertTrue(wrote_decision)
        self.assertTrue(wrote_ledger)

    def test_annotation_scout_rejects_cross_genus_transcript_first(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            query = tmp_path / "queries.faa"
            self.write_query_fasta(query)
            decision, _rows = genecluster_annotation_scout.build_route_decision(
                query,
                [
                    {
                        "source_id": "mismatch-rna",
                        "organism": "Coptis chinensis",
                        "transcriptome": "provider://rna/srr",
                        "transcriptome_species": "Argemone mexicana",
                    }
                ],
            )

        rejected = {record["route"]: record["reason"] for record in decision["rejected_routes"]}
        self.assertEqual("next_experiment_design", decision["selected_route"])
        self.assertIn("transcriptome_species_mismatch", decision["blockers"])
        self.assertIn("transcript_first", rejected)
        self.assertIn("does not match target genus", rejected["transcript_first"])

    def test_annotation_scout_blocks_queries_without_required_controls(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            query = tmp_path / "queries.faa"
            self.write_query_fasta(query, include_controls=False)
            proteome, gff = self.write_annotation_pair(tmp_path)
            decision, _rows = genecluster_annotation_scout.build_route_decision(
                query,
                [{"source_id": "joined", "organism": "Coptis chinensis", "proteome_fasta": str(proteome), "gff": str(gff)}],
            )

        self.assertEqual("next_experiment_design", decision["selected_route"])
        self.assertFalse(decision["controls"]["ok"])
        self.assertEqual(["ACT2", "GAPDH", "random_shuffle"], decision["controls"]["missing_controls"])
        self.assertIn("missing_required_query_controls", decision["blockers"])

    def test_query_registry_marks_intake_blocked_query_with_correct_ceiling(self) -> None:
        registry = SKILL_ROOT / "references" / "genecluster-query-registry.tsv"
        required_claims = SKILL_ROOT / "references" / "genecluster-required-claims.tsv"
        registry_result = genecluster_preflight.validate_query_registry(registry)
        claims_result = genecluster_preflight.validate_required_claims(required_claims, query_registry=registry)

        self.assertEqual([], registry_result["errors"])
        self.assertEqual([], claims_result["errors"])
        rows, _fields = genecluster_preflight.read_tsv(registry)
        intake_blocked = next(row for row in rows if row["sequence_status"] == "unresolved_intake_blocked")
        self.assertEqual("not_tested_intake_blocked", intake_blocked["claim_ceiling_if_unresolved"])
        self.assertNotIn("absent", intake_blocked["claim_ceiling_if_unresolved"])

    def test_required_claim_gate_blocks_asserted_claim_when_query_unresolved(self) -> None:
        registry = SKILL_ROOT / "references" / "genecluster-query-registry.tsv"
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "required-claims.tsv"
            self.write_tsv(
                path,
                list(genecluster_preflight.REQUIRED_CLAIM_COLUMNS),
                [
                    {
                        "claim_id": "CLAIM-EXAMPLE-INTAKE-BLOCKED",
                        "statement": "Assert example pathway gene search result.",
                        "required_query_ids": "Q-EXAMPLE-001",
                        "required_source_classes": "protein_seed",
                        "minimum_query_status": "resolved",
                        "allowed_if_unresolved": "false",
                        "blocked_output_labels": "absent;0 hits;missing from target;no homolog found",
                        "review_gate": "claim_audit",
                        "assertion_status": "asserted",
                    }
                ],
            )
            result = genecluster_preflight.validate_required_claims(path, query_registry=registry)

        self.assertFalse(result["ok"])
        self.assertTrue(any("unresolved required queries" in error for error in result["errors"]))

    def test_source_scout_writes_deterministic_no_network_ledgers(self) -> None:
        registry = SKILL_ROOT / "references" / "genecluster-query-registry.tsv"
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "source-scout"
            report, records, source_rows = genecluster_source_scout.build_report(registry, None, out)
            genecluster_source_scout.write_outputs(report, records, source_rows, out)
            preflight_result = genecluster_preflight.validate_source_ledger(out / "source-ledger.tsv")

        self.assertTrue(report["ok"])
        self.assertEqual("unresolved_intake_blocked", records[0]["resolution_status"])
        self.assertTrue(any(row["source_name"] == "NGDC GWH/CNCB" for row in source_rows))
        self.assertTrue(all(row["source_record_type"] == "source_scout_probe" for row in source_rows))
        self.assertTrue(all(row["source_provider"] for row in source_rows))
        self.assertTrue(all(row["acquisition_policy"] == "metadata_only_no_network_no_raw_download" for row in source_rows))
        self.assertTrue(all(row["network_call_planned"] == "false" for row in source_rows))
        self.assertEqual([], preflight_result["errors"])

    def test_source_ledger_rejects_network_call_in_mixed_source_scout_ledger(self) -> None:
        registry = SKILL_ROOT / "references" / "genecluster-query-registry.tsv"
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "source-scout"
            report, records, source_rows = genecluster_source_scout.build_report(registry, None, out)
            source_rows[0]["network_call_planned"] = "true"
            genecluster_source_scout.write_outputs(report, records, source_rows, out)
            result = genecluster_preflight.validate_source_ledger(out / "source-ledger.tsv")
        self.assertFalse(result["ok"])
        self.assertTrue(any("must not plan network calls" in error for error in result["errors"]))

    def test_source_scout_continues_after_ncbi_miss_to_ngdc_hit(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            registry = tmp_path / "query-registry.tsv"
            registry.write_text(
                "query_id\tquery_name\tsource_organism\tclaim_id\tclaim_ceiling\tresolution_status\t"
                "ncbi_datasets_api_status\tncbi_protein_status\tuniprot_rest_status\tngdc_gwh_cncb\tngdc_gwh_cncb_status\n"
                "Q001\tlate source candidate\tArabidopsis thaliana\tCLAIM_Q001\t\t\t"
                "not_found\tnot_found\tnot_found\tGWHATEST00000\tcandidate_source_available\n",
                encoding="utf-8",
            )
            report, records, _source_rows = genecluster_source_scout.build_report(registry, None, tmp_path / "source-scout")

        self.assertTrue(report["ok"])
        self.assertEqual("registry_reference_present", records[0]["resolution_status"])
        self.assertEqual("NGDC GWH/CNCB", records[0]["selected_source"])
        self.assertEqual("GWHATEST00000", records[0]["selected_reference"])

    def test_public_fixture_route_scout_selects_annotation_direct(self) -> None:
        fixtures = EXAMPLE / "fixtures"
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "route-scout"
            sources = genecluster_annotation_scout.load_source_ledger(fixtures / "route-source-ledger.tsv")
            decision, rows = genecluster_annotation_scout.build_route_decision(
                fixtures / "query-with-controls.faa",
                sources,
                campaign=json.loads((EXAMPLE / "campaign-manifest.json").read_text(encoding="utf-8")),
            )
            outputs = genecluster_annotation_scout.write_route_outputs(decision, rows, out)
            preflight_result = genecluster_preflight.validate_route_annotation_ledger(
                Path(outputs["annotation_ledger"]),
                repo_root=ROOT,
            )

        self.assertEqual("annotation_direct", decision["selected_route"])
        self.assertEqual("L3_annotation_neighborhood_ready", decision["claim_ceiling"])
        self.assertEqual("example_fixture", decision["selected_source_id"])
        self.assertEqual(2, rows[0]["protein_gff_join_count"])
        self.assertEqual([], preflight_result["errors"])

    def test_annotation_scout_joins_ngdc_protein_accession_gff(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            query = tmp_path / "queries.faa"
            self.write_query_fasta(query)
            proteome = tmp_path / "proteome.faa"
            proteome.write_text(">GWHPTEST001 annotated protein\nMAAAAA\n", encoding="utf-8")
            gff = tmp_path / "genomic.gff"
            gff.write_text(
                "chr1\tNGDC\tCDS\t10\t99\t.\t+\t0\tID=cds1;Parent=rna-1;Protein_Accession=GWHPTEST001\n",
                encoding="utf-8",
            )
            decision, rows = genecluster_annotation_scout.build_route_decision(
                query,
                [{"source_id": "ngdc", "organism": "Arabidopsis thaliana", "proteome_fasta": str(proteome), "gff": str(gff)}],
            )

        self.assertEqual("annotation_direct", decision["selected_route"])
        self.assertEqual(1, rows[0]["protein_gff_join_count"])

    def test_bgc_consensus_validator_rejects_collapsed_caller_disagreement(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            calls = tmp_path / "cluster_calls.tsv"
            consensus = tmp_path / "bgc_consensus.tsv"
            self.write_tsv(
                calls,
                list(genecluster_atlas_contracts.CLUSTER_CALL_COLUMNS),
                [
                    {
                        "cluster_id": "cluster-1",
                        "caller": "plantiSMASH",
                        "source_species": "Coptis chinensis",
                        "target_species": "Coptis chinensis",
                        "contig": "chr1",
                        "start": "10",
                        "end": "100",
                        "core_genes": "protA",
                        "confidence": "0.8",
                        "claim_level": "L4_consensus_supported",
                    },
                    {
                        "cluster_id": "cluster-1",
                        "caller": "cblaster",
                        "source_species": "Coptis chinensis",
                        "target_species": "Coptis chinensis",
                        "contig": "chr1",
                        "start": "10",
                        "end": "150",
                        "core_genes": "protA,protB",
                        "confidence": "0.6",
                        "claim_level": "L4_consensus_supported",
                    },
                ],
            )
            self.write_tsv(
                consensus,
                list(genecluster_atlas_contracts.BGC_CONSENSUS_COLUMNS),
                [
                    {
                        "consensus_id": "consensus-1",
                        "cluster_id": "cluster-1",
                        "verdict": "supported",
                        "caller_count": "2",
                        "agreeing_callers": "plantiSMASH,cblaster",
                        "disagreeing_callers": "",
                        "disagreement_status": "none",
                        "claim_level": "L4_consensus_supported",
                        "caller_versions": "plantiSMASH=fixture;cblaster=fixture",
                        "caller_licenses": "plantiSMASH=academic;cblaster=MIT",
                    }
                ],
            )
            result = genecluster_atlas_contracts.validate_bgc_consensus(consensus, cluster_calls=calls)

        self.assertFalse(result["ok"])
        self.assertTrue(any("collapses caller disagreement" in error for error in result["errors"]))

    def test_bgc_consensus_allows_identical_boundaries_from_multiple_callers(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            calls = tmp_path / "cluster_calls.tsv"
            consensus = tmp_path / "bgc_consensus.tsv"
            rows = [
                {
                    "cluster_id": "cluster-1",
                    "caller": "plantiSMASH",
                    "source_species": "Coptis chinensis",
                    "target_species": "Coptis chinensis",
                    "contig": "chr1",
                    "start": "10",
                    "end": "100",
                    "core_genes": "protA;protB",
                    "confidence": "0.8",
                    "claim_level": "L4_consensus_supported",
                },
                {
                    "cluster_id": "cluster-1",
                    "caller": "cblaster",
                    "source_species": "Coptis chinensis",
                    "target_species": "Coptis chinensis",
                    "contig": "chr1",
                    "start": "10",
                    "end": "100",
                    "core_genes": "protB;protA",
                    "confidence": "0.7",
                    "claim_level": "L4_consensus_supported",
                },
            ]
            self.write_tsv(calls, list(genecluster_atlas_contracts.CLUSTER_CALL_COLUMNS), rows)
            self.write_tsv(
                consensus,
                list(genecluster_atlas_contracts.BGC_CONSENSUS_COLUMNS),
                [
                    {
                        "consensus_id": "consensus-1",
                        "cluster_id": "cluster-1",
                        "verdict": "supported",
                        "caller_count": "2",
                        "agreeing_callers": "plantiSMASH,cblaster",
                        "disagreeing_callers": "",
                        "disagreement_status": "none",
                        "claim_level": "L4_consensus_supported",
                        "caller_versions": "plantiSMASH=fixture;cblaster=fixture",
                        "caller_licenses": "plantiSMASH=review;cblaster=MIT",
                    }
                ],
            )
            result = genecluster_atlas_contracts.validate_bgc_consensus(consensus, cluster_calls=calls)

        self.assertEqual([], result["errors"])

    def test_function_jury_validator_rejects_collapsed_contradictions(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            votes = tmp_path / "protein_function_votes.tsv"
            jury = tmp_path / "protein_function_jury.tsv"
            self.write_tsv(
                votes,
                list(genecluster_atlas_contracts.PROTEIN_FUNCTION_VOTE_COLUMNS),
                [
                    {
                        "protein_id": "protA",
                        "tool": "Pfam",
                        "function_label": "methyltransferase",
                        "confidence": "0.7",
                        "evidence_level": "L3_annotation_neighborhood_ready",
                        "tool_version": "fixture",
                        "license": "open",
                    },
                    {
                        "protein_id": "protA",
                        "tool": "DIAMOND",
                        "function_label": "transporter",
                        "confidence": "0.6",
                        "evidence_level": "L3_annotation_neighborhood_ready",
                        "tool_version": "fixture",
                        "license": "open",
                    },
                ],
            )
            self.write_tsv(
                jury,
                list(genecluster_atlas_contracts.PROTEIN_FUNCTION_JURY_COLUMNS),
                [
                    {
                        "protein_id": "protA",
                        "verdict": "methyltransferase",
                        "claim_level": "L4_consensus_supported",
                        "supporting_tools": "Pfam",
                        "contradicting_tools": "",
                        "confidence": "0.65",
                    }
                ],
            )
            result = genecluster_atlas_contracts.validate_protein_function_jury(jury, votes=votes)

        self.assertFalse(result["ok"])
        self.assertTrue(any("collapses contradictory function votes" in error for error in result["errors"]))

    def test_comparative_atlas_fixture_validates(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            atlas = Path(tmp) / "comparative_atlas"
            atlas.mkdir()
            self.write_tsv(
                atlas / "species-ledger.tsv",
                ["species_id", "scientific_name", "assembly_id", "annotation_id", "data_status", "license"],
                [{"species_id": "sp1", "scientific_name": "Coptis chinensis", "assembly_id": "GCA_fixture", "annotation_id": "ann1", "data_status": "fixture", "license": "open"}],
            )
            self.write_tsv(
                atlas / "orthogroups.tsv",
                ["orthogroup_id", "species_id", "protein_id", "paralog_group", "orthology_status"],
                [{"orthogroup_id": "OG1", "species_id": "sp1", "protein_id": "protA", "paralog_group": "P1", "orthology_status": "ortholog"}],
            )
            self.write_tsv(
                atlas / "synteny_blocks.tsv",
                ["block_id", "species_id", "contig", "start", "end", "anchor_gene", "support_status"],
                [{"block_id": "B1", "species_id": "sp1", "contig": "chr1", "start": "1", "end": "1000", "anchor_gene": "protA", "support_status": "fixture"}],
            )
            self.write_tsv(
                atlas / "cluster_call_matrix.tsv",
                ["cluster_id", "species_id", "caller", "call_status", "claim_level"],
                [{"cluster_id": "cluster-1", "species_id": "sp1", "caller": "plantiSMASH", "call_status": "present", "claim_level": "L4_consensus_supported"}],
            )
            self.write_tsv(
                atlas / "comparative_neighborhoods.tsv",
                ["neighborhood_id", "species_id", "cluster_id", "gene_id", "relative_order", "function_label"],
                [{"neighborhood_id": "N1", "species_id": "sp1", "cluster_id": "cluster-1", "gene_id": "protA", "relative_order": "0", "function_label": "anchor"}],
            )
            (atlas / "atlas-summary.md").write_text("# Atlas Summary\n\nFixture summary.\n", encoding="utf-8")
            result = genecluster_atlas_contracts.validate_comparative_atlas(atlas)

        self.assertEqual([], result["errors"])

    def test_atlas_normalizers_emit_contract_valid_fixture_outputs(self) -> None:
        cluster_rows = [
            {"cluster_id": "cluster-1", "contig": "chr1", "start": "10", "end": "100", "core_genes": "protA;protB", "confidence": "0.8"},
        ]
        plant_rows = [
            {"cluster_id": "cluster-1", "contig": "chr1", "start": "10", "end": "100", "core_genes": "protA;protB", "confidence": "0.7"},
        ]
        pfam_rows = [{"protein_id": "protA", "pfam_name": "methyltransferase", "evalue": "1e-20"}]
        clean_rows = [{"protein_id": "protA", "ec_number": "2.1.1.1", "confidence": "0.8"}]
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            cluster_calls = genecluster_atlas_normalizers.normalize_cluster_calls(
                annotation_direct_rows=cluster_rows,
                plantismash_rows=plant_rows,
                source_species="Coptis chinensis",
                target_species="Coptis chinensis",
            )
            consensus = genecluster_atlas_normalizers.build_bgc_consensus(cluster_calls)
            votes = genecluster_atlas_normalizers.normalize_protein_function_votes(
                pfam_rows=pfam_rows,
                clean_rows=clean_rows,
            )
            jury = genecluster_atlas_normalizers.build_protein_function_jury(votes)
            genecluster_atlas_normalizers.write_tsv(tmp_path / "cluster_calls.tsv", genecluster_atlas_normalizers.CLUSTER_CALL_COLUMNS, cluster_calls)
            genecluster_atlas_normalizers.write_tsv(tmp_path / "bgc_consensus.tsv", genecluster_atlas_normalizers.BGC_CONSENSUS_COLUMNS, consensus)
            genecluster_atlas_normalizers.write_tsv(tmp_path / "protein_function_votes.tsv", genecluster_atlas_normalizers.PROTEIN_FUNCTION_VOTE_COLUMNS, votes)
            genecluster_atlas_normalizers.write_tsv(tmp_path / "protein_function_jury.tsv", genecluster_atlas_normalizers.PROTEIN_FUNCTION_JURY_COLUMNS, jury)
            atlas = genecluster_atlas_normalizers.build_comparative_atlas(
                species_rows=[{"species_id": "sp1", "scientific_name": "Coptis chinensis", "assembly_id": "GCA_fixture", "annotation_id": "ann1", "data_status": "fixture", "license": "open"}],
                orthofinder_rows=[{"orthogroup_id": "OG1", "species_id": "sp1", "protein_id": "protA", "paralog_group": "P1", "orthology_status": "ortholog"}],
                genespace_rows=[{"block_id": "B1", "species_id": "sp1", "contig": "chr1", "start": "10", "end": "100", "anchor_gene": "protA", "support_status": "supported"}],
                cluster_call_rows=cluster_calls,
                protein_jury_rows=jury,
            )
            genecluster_atlas_normalizers.write_comparative_atlas(tmp_path / "comparative_atlas", atlas)

            checks = [
                genecluster_atlas_contracts.validate_cluster_calls(tmp_path / "cluster_calls.tsv"),
                genecluster_atlas_contracts.validate_bgc_consensus(tmp_path / "bgc_consensus.tsv", cluster_calls=tmp_path / "cluster_calls.tsv"),
                genecluster_atlas_contracts.validate_protein_function_votes(tmp_path / "protein_function_votes.tsv"),
                genecluster_atlas_contracts.validate_protein_function_jury(tmp_path / "protein_function_jury.tsv", votes=tmp_path / "protein_function_votes.tsv"),
                genecluster_atlas_contracts.validate_comparative_atlas(tmp_path / "comparative_atlas"),
            ]

        errors = [error for result in checks for error in result["errors"]]
        self.assertEqual([], errors)

    def test_review_and_provider_manifests_reject_raw_artifacts_and_secret_values(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            review = tmp_path / "review_surface_manifest.json"
            review.write_text(
                json.dumps(
                    {
                        "schema_version": "genecluster_review_surface.v1",
                        "review_id": "review-1",
                        "source_tables": [{"path": "cluster_neighborhoods.tsv", "artifact_type": "summary_table"}],
                        "generated_files": [{"path": "review/raw_reads.fastq", "artifact_type": "raw_reads"}],
                        "claims": [
                            {
                                "claim_id": "claim-1",
                                "statement": "Fixture claim",
                                "claim_level": "L4_consensus_supported",
                                "evidence_level": "L4_consensus_supported",
                                "caveat": "Fixture only",
                                "review_status": "accepted",
                            }
                        ],
                    }
                ),
                encoding="utf-8",
            )
            provider = tmp_path / "provider_handoff_manifest.json"
            provider.write_text(
                json.dumps(
                    {
                        "schema_version": "genecluster_provider_handoff.v1",
                        "provider": {"adapter": "runpod_bridge", "mutation_owner": "host_side_hook"},
                        "workload": {"stage_contract": "stage-contract.json", "route_decision": "route_decision.json"},
                        "artifact_egress": {"expected_artifacts": [{"path": "run-summary.json", "artifact_type": "summary_json"}]},
                        "safety": {
                            "credentials": [{"env_name": "RUNPOD_API_KEY", "source": "env"}],
                            "runpod_api_key": "literal_secret_value_for_test",
                        },
                    }
                ),
                encoding="utf-8",
            )

            review_result = genecluster_atlas_contracts.validate_review_surface_manifest(review)
            provider_result = genecluster_atlas_contracts.validate_provider_handoff_manifest(provider)

        self.assertFalse(review_result["ok"])
        self.assertTrue(any("raw/heavy" in error or "artifact_type is forbidden" in error for error in review_result["errors"]))
        self.assertFalse(provider_result["ok"])
        self.assertTrue(any("secret" in error or "literal value" in error for error in provider_result["errors"]))

    def test_provider_manifest_allows_env_names_and_summary_artifacts(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            provider = Path(tmp) / "provider_handoff_manifest.json"
            provider.write_text(
                json.dumps(
                    {
                        "schema_version": "genecluster_provider_handoff.v1",
                        "provider": {"adapter": "runpod_bridge", "mutation_owner": "host_side_hook"},
                        "workload": {
                            "stage_contract": "stage-contract.json",
                            "route_decision": "route_decision.json",
                            "cost_bounds": {"max_usd": "1.00", "stop_when_exceeded": "stop_provider_run"},
                        },
                        "artifact_egress": {
                            "summary_only": True,
                            "hash_algorithm": "sha256",
                            "hash_ledger": "hashes.tsv",
                            "expected_artifacts": [{"path": "run-summary.json", "artifact_type": "summary_json"}],
                        },
                        "cleanup": {"verify_pod_stopped": True, "verify_artifacts_fetched": True},
                        "safety": {"credentials": [{"env_name": "RUNPOD_API_KEY", "source": "env"}]},
                    }
                ),
                encoding="utf-8",
            )
            result = genecluster_atlas_contracts.validate_provider_handoff_manifest(provider)

        self.assertEqual([], result["errors"])

    def test_static_review_surface_validates_and_omits_raw_artifacts(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            final = tmp_path / "final"
            final.mkdir()
            (final / "biology-coptis-chinensis.md").write_text("# Coptis\n\nAnnotation summary.\n", encoding="utf-8")
            (final / "coptis-chinensis-bia-pathway-2026-05-07.xlsx").write_text("fixture workbook placeholder", encoding="utf-8")
            out = tmp_path / "review"
            build_result = genecluster_review_surface.build_review_surface(
                out,
                review_id="fixture-review",
                final_deliverable=final,
            )
            manifest = Path(build_result["manifest"])
            validate_result = genecluster_atlas_contracts.validate_review_surface_manifest(manifest)
            manifest_text = manifest.read_text(encoding="utf-8")

        self.assertEqual([], validate_result["errors"])
        self.assertIn("index.html", manifest_text)
        self.assertNotIn(".faa", manifest_text)
        self.assertNotIn(".gff", manifest_text)


if __name__ == "__main__":
    unittest.main()
