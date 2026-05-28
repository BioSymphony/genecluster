#!/usr/bin/env python3
"""Convert old GeneCluster request workbooks into control-plane ledgers.

The intake is deliberately small-file only. It reads an operator-provided Excel
request sheet, extracts dataset links, goal notes, and embedded protein query
sequences, then writes GeneCluster manifests/ledgers that can be validated and
used to generate launch bundles. It does not fetch SRA records, resolve protein
accessions, or run search tools.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import sys
from datetime import date
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except Exception as exc:  # pragma: no cover - exercised only on missing envs
    raise SystemExit(
        "openpyxl is required for GeneCluster Excel intake. "
        "Install workspace dependencies or run with the bundled Codex Python."
    ) from exc


TODAY = date.today().isoformat()
REMOTE_RUN_ROOT = "/workspace/genecluster/runs/<run_id>"
DB_CACHE_ROOT = "/workspace/genecluster/db-cache"


def safe_id(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9]+", "_", value.strip().lower()).strip("_")
    return cleaned or "unknown"


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", " ").strip()


def clean_sequence(value: Any) -> str:
    text = clean_text(value)
    text = re.sub(r"^<-\s*", "", text)
    text = re.sub(r"[^A-Za-z*]", "", text).upper()
    if not text or "YOUWILLHAVETOFIND" in text:
        return ""
    return text


def extract_accession(url: str) -> str:
    for pattern in (r"(SRX\d+)", r"(SRR\d+)", r"(PRJ[A-Z]+\d+)", r"(SAM[NED][A-Z]?\d+)"):
        match = re.search(pattern, url)
        if match:
            return match.group(1)
    uid = re.search(r"from_uid=(\d+)", url)
    if uid:
        return f"BIOPROJECT_UID_{uid.group(1)}"
    return "remote_resolve_required"


def infer_target_from_section(section: str) -> str:
    section_l = section.lower()
    if "uncaria" in section_l:
        return "Uncaria sp."
    if "Coptis" in section_l:
        return "Coptis chinensis"
    return "target species"


def infer_data_role(label: str) -> str:
    label_l = label.lower()
    if "genome" in label_l:
        return "genome_wgs"
    if "iso" in label_l or "long" in label_l:
        return "transcriptome_isoseq"
    if "transcriptome" in label_l:
        return "transcriptome_rna"
    return "sequence_resource"


def infer_sample_type(label: str, notes: str) -> str:
    combined = f"{label} {notes}".lower()
    if "capsule" in combined:
        return "capsule tissue; mid-stage; from workbook note"
    if "genome" in combined:
        return "gDNA; metadata remote_resolve_required"
    if "transcriptome" in combined:
        return "RNA-seq; metadata remote_resolve_required"
    return "metadata remote_resolve_required"


def infer_source_organism(name: str, description: str) -> str:
    combined = f"{name} {description}".lower()
    if "catharanthus" in combined or "c.roseus" in combined or "c roseus" in combined:
        return "Catharanthus roseus"
    if "vinca minor" in combined or name.lower().startswith("vm"):
        return "Vinca minor"
    if "coffea arabica" in combined or name.lower().startswith("ca"):
        return "Coffea arabica"
    if "uncaria" in combined or name.lower().startswith("um"):
        return "Uncaria sp."
    if "mitspec" in combined or name.lower().startswith("ms"):
        return "Coptis chinensis"
    return "source organism unresolved from workbook"


def infer_enzyme_class(name: str, description: str) -> str:
    combined = f"{name} {description}".lower()
    if "strictosidine synthase" in combined or "str1" in combined:
        return "strictosidine synthase"
    if "strictosidine glucosidase" in combined or "sgd" in combined:
        return "beta-glucosidase"
    if "cyp" in combined or "cytochrome" in combined:
        return "cytochrome P450"
    if "methyltransf" in combined or "methyltransferase" in combined or "omt" in combined or "lamt" in combined or "mt" in name.lower():
        return "methyltransferase"
    if "dcs" in combined or "pln02586" in combined or "dehydrogenase" in combined or "reductase" in combined:
        return "medium-chain dehydrogenase/reductase"
    if "transporter" in combined or "npf" in combined or "mate" in combined:
        return "transporter"
    return "enzyme class unresolved"


def infer_pathway_role(name: str, description: str) -> str:
    enzyme = infer_enzyme_class(name, description)
    if enzyme == "strictosidine synthase":
        return "strictosidine formation anchor"
    if enzyme == "beta-glucosidase":
        return "strictosidine aglycone formation anchor"
    if enzyme == "medium-chain dehydrogenase/reductase":
        return "post-strictosidine reductase / corynanthe branch candidate"
    if enzyme == "methyltransferase":
        return "methyltransferase tailoring candidate"
    if enzyme == "cytochrome P450":
        return "hydroxylation / oxidative tailoring candidate"
    if enzyme == "transporter":
        return "transport or neighborhood context"
    return "candidate pathway query from workbook"


def infer_family_scope(enzyme_class: str, description: str) -> str:
    if description:
        return re.sub(r"\s+", "_", description.strip().replace("[", "").replace("]", "").replace("'", ""))[:120]
    return safe_id(enzyme_class)


def motif_requirements(enzyme_class: str) -> str:
    return {
        "strictosidine synthase": "strictosidine synthase fold/domain review; catalytic residue review",
        "beta-glucosidase": "GH1 beta-glucosidase motifs; strictosidine-specific clade review",
        "cytochrome P450": "P450 heme-binding motif and PERF motif; family/clade placement required",
        "methyltransferase": "methyltransferase/SABATH motifs; substrate-specificity caveat",
        "medium-chain dehydrogenase/reductase": "MDR/ADH_N architecture; NAD(P)-binding and catalytic tetrad review",
        "transporter": "transporter domains; context-only unless orthogonal evidence supports pathway role",
    }.get(enzyme_class, "domain architecture review required")


def false_positive_risk(enzyme_class: str, has_sequence: bool) -> tuple[str, str, str]:
    broad = enzyme_class in {"cytochrome P450", "methyltransferase", "beta-glucosidase", "transporter"}
    if not has_sequence:
        return "true", "high", "sequence_resolver_required"
    if broad:
        return "true", "high", f"broad_{safe_id(enzyme_class)}_decoys_required"
    return "false", "medium", f"{safe_id(enzyme_class)}_family_decoys_required"


def read_workbook(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, data_only=False, read_only=False)
    ws = wb.active
    goals: list[str] = []
    datasets: list[dict[str, str]] = []
    queries: list[dict[str, str]] = []
    current_section = ""
    uncaria_note = ""
    in_query_table = False

    for row in ws.iter_rows(values_only=True):
        cells = [clean_text(cell) for cell in row]
        row_text = " ".join(cell for cell in cells if cell)
        if not row_text:
            continue
        if "Links to Coptis sequence data" in row_text:
            current_section = "Coptis sequence data"
            continue
        if "Link to Uncaria sequence data" in row_text:
            current_section = "Uncaria sequence data"
            continue
        if row_text.startswith("^(capsule"):
            uncaria_note = row_text
            continue
        if "Protein queries" in row_text:
            in_query_table = True
            continue
        if in_query_table and cells[:5]:
            if any(cell == "Name" for cell in cells):
                continue
            number = cells[1] if len(cells) > 1 else ""
            name = cells[2] if len(cells) > 2 else ""
            description = cells[3] if len(cells) > 3 else ""
            sequence = cells[4] if len(cells) > 4 else ""
            if number.isdigit() and name:
                queries.append(
                    {
                        "number": number,
                        "name": name,
                        "description": description,
                        "sequence": clean_sequence(sequence),
                        "sequence_note": sequence if not clean_sequence(sequence) else "",
                    }
                )
                continue
        if cells and cells[0].startswith("Find "):
            goals.append(cells[0])
        if len(cells) > 4 and cells[4].startswith("Find "):
            goals.append(cells[4])
        if len(cells) > 2 and cells[2].startswith("https://"):
            label = cells[1] or "sequence resource"
            url = cells[2]
            section_species = infer_target_from_section(current_section)
            note = uncaria_note if "Uncaria" in current_section else ""
            datasets.append(
                {
                    "label": label,
                    "url": url,
                    "section": current_section or "unspecified sequence data",
                    "organism": section_species,
                    "data_role": infer_data_role(label),
                    "sample_type": infer_sample_type(label, note),
                    "note": note,
                    "accession": extract_accession(url),
                }
            )

    return {"sheet": ws.title, "datasets": datasets, "queries": queries, "goals": goals}


def write_tsv(path: Path, columns: list[str], rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, delimiter="\t", extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({col: row.get(col, "") for col in columns})


def write_json(path: Path, data: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, indent=2, sort_keys=False) + "\n", encoding="utf-8")


def build_data_ledger(parsed: dict[str, Any], campaign_id: str) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    seen: dict[str, int] = {}
    for item in parsed["datasets"]:
        organism_id = safe_id(item["organism"])
        role_id = safe_id(item["data_role"])
        seen_key = f"{organism_id}_{role_id}"
        seen[seen_key] = seen.get(seen_key, 0) + 1
        dataset_id = f"{seen_key}_{seen[seen_key]}"
        rows.append(
            {
                "dataset_id": dataset_id,
                "accession": item["accession"],
                "run_id": "remote_resolve_required",
                "data_role": item["data_role"],
                "sample_type": item["sample_type"],
                "organism": item["organism"],
                "bioproject": "remote_resolve_required",
                "technology": "remote_resolve_required",
                "expected_size": "remote_resolve_required",
                "source_url": item["url"],
                "remote_path": f"{REMOTE_RUN_ROOT}/inputs/{dataset_id}",
                "checksum_status": "remote_pending",
                "data_sensitivity": "public",
                "allowed_compute_location": "runpod_or_user_configured_remote",
                "allowed_upload": "no_public_webserver_upload",
                "redistribution_policy": "derived-summary-only",
                "terms_checked_date": TODAY,
                "license_url": "https://www.ncbi.nlm.nih.gov/home/about/policies/",
                "citation_doi": "remote_resolve_required",
                "md5_or_sha256": "remote_pending",
                "frozen_metadata_path": "data/data-ledger.tsv",
                "raw_artifact_policy": "remote_only",
                "retention_policy": "provider_volume_until_review",
                "operator_approval_id": "operator_review_required",
                "notes": f"Imported from workbook for {campaign_id}; {item['section']}; {item.get('note', '')}".strip(),
            }
        )
    return rows


def build_query_ledger(parsed: dict[str, Any]) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    rows: list[dict[str, str]] = []
    fasta_records: list[dict[str, str]] = []
    for query in parsed["queries"]:
        query_id = f"Q{int(query['number']):03d}"
        name = query["name"]
        description = query["description"]
        seq = query["sequence"]
        enzyme = infer_enzyme_class(name, description)
        source_organism = infer_source_organism(name, description)
        broad_flag, risk, negatives = false_positive_risk(enzyme, bool(seq))
        checksum = hashlib.sha256(seq.encode("utf-8")).hexdigest() if seq else "remote_pending"
        sequence_source = "workbook_embedded_protein_sequence" if seq else "workbook_resolver_note"
        sequence_type = "protein_seed" if seq else "literature_seed"
        curation_status = "resolved" if seq else "remote_resolve_required"
        confidence = "medium" if seq else "context"
        if enzyme in {"strictosidine synthase", "beta-glucosidase", "medium-chain dehydrogenase/reductase"} and seq:
            confidence = "high"
        row = {
            "query_id": query_id,
            "query_name": name,
            "source_organism": source_organism,
            "sequence_source": sequence_source,
            "enzyme_class": enzyme,
            "pathway_role": infer_pathway_role(name, description),
            "confidence": confidence,
            "citation": "",
            "resolved_accession": "workbook_sequence" if seq else "remote_resolve_required",
            "sequence_type": sequence_type,
            "sequence_length": str(len(seq)) if seq else "remote_pending",
            "checksum": checksum,
            "family_scope": infer_family_scope(enzyme, description),
            "motif_requirements": motif_requirements(enzyme),
            "negative_controls": negatives,
            "decoy_or_broad_family_flag": broad_flag,
            "expected_false_positive_risk": risk,
            "curation_status": curation_status,
            "last_resolved_at": TODAY,
            "notes": description or query.get("sequence_note", ""),
        }
        rows.append(row)
        if seq:
            fasta_records.append({"query_id": query_id, "name": name, "sequence": seq})
    return rows, fasta_records


def build_resource_ledger() -> list[dict[str, str]]:
    return [
        {
            "resource": "BLAST+",
            "resource_type": "tool",
            "version": "provider_toolcheck",
            "license_class": "academic-free-or-web",
            "use_mode": "provider-local search only",
            "citation": "https://www.ncbi.nlm.nih.gov/books/NBK279690/",
        },
        {
            "resource": "DIAMOND",
            "resource_type": "tool",
            "version": "provider_toolcheck",
            "license_class": "permissive-code",
            "use_mode": "provider-local protein search",
            "citation": "https://github.com/bbuchfink/diamond",
        },
        {
            "resource": "MMseqs2",
            "resource_type": "tool",
            "version": "provider_toolcheck",
            "license_class": "permissive-code",
            "use_mode": "provider-local clustering/search",
            "citation": "https://github.com/soedinglab/MMseqs2",
        },
        {
            "resource": "Pfam/HMMER",
            "resource_type": "database_and_tool",
            "version": "provider_cache_or_remote_resolve",
            "license_class": "open-data-with-terms",
            "use_mode": "provider-local domain scan",
            "citation": "https://pfam.xfam.org/",
        },
    ]


def build_database_ledger() -> list[dict[str, str]]:
    records = [
        ("blast_swissprot", "blast", "protein", f"{DB_CACHE_ROOT}/blast/swissprot", "remote_preload_required", "NCBI BLAST swissprot", "false", "required", "candidate_search", "small", "high", "download_preformatted_blastdb"),
        ("diamond_swissprot", "diamond", "protein", f"{DB_CACHE_ROOT}/diamond/swissprot.dmnd", "remote_preload_required", "UniProt Swiss-Prot", "false", "required", "candidate_search", "small", "high", "build_or_preload_provider_side"),
        ("mmseqs_uniprotkb", "mmseqs", "protein", f"{DB_CACHE_ROOT}/mmseqs/uniprotkb", "remote_preload_required", "UniProtKB", "false", "required", "candidate_search", "medium", "high", "createindex_provider_side"),
        ("hmmer_pfam", "hmmer", "domain", f"{DB_CACHE_ROOT}/hmmer/Pfam-A.hmm", "remote_preload_required", "Pfam-A HMM", "false", "required", "candidate_search", "medium", "high", "preload_hmm_and_run_hmmpress"),
        ("custom_workbook_query_proteins", "custom", "protein", f"{REMOTE_RUN_ROOT}/inputs/query-sequences.faa", "workbook_import", "workbook", "false", "optional", "candidate_search", "tiny", "high", "copy_small_query_fasta_to_provider_inputs"),
        ("custom_workbook_query_blast", "blast", "protein", f"{DB_CACHE_ROOT}/custom/workbook-query-proteins-blast", "workbook_import", "workbook", "true", "optional", "candidate_search", "tiny", "high", "makeblastdb_provider_side"),
        ("custom_workbook_query_diamond", "diamond", "protein", f"{DB_CACHE_ROOT}/custom/workbook-query-proteins.dmnd", "workbook_import", "workbook", "true", "optional", "candidate_search", "tiny", "high", "diamond_makedb_provider_side"),
        ("custom_target_transcripts", "blast", "nucleotide", f"{REMOTE_RUN_ROOT}/databases/target-transcripts", "remote_run", "provider_target_builder", "true", "optional", "candidate_search", "medium", "high", "build_from_provider_materialized_target_transcripts"),
        ("custom_target_proteins", "diamond", "protein", f"{REMOTE_RUN_ROOT}/databases/target-proteins.dmnd", "remote_run", "provider_target_builder", "true", "optional", "candidate_search", "medium", "high", "build_from_provider_materialized_target_proteins"),
        ("cdd_rpsblast", "rpsblast", "domain", f"{DB_CACHE_ROOT}/cdd/Cdd", "remote_preload_required", "NCBI CDD", "false", "optional", "full_public_mining", "medium", "medium", "download_preformatted_cdd"),
    ]
    rows = []
    for db_id, engine, seq_type, remote_path, version, source, build_required, priority, run_gate, cost_class, prep_roi, bootstrap_strategy in records:
        rows.append(
            {
                "db_id": db_id,
                "engine": engine,
                "sequence_type": seq_type,
                "remote_path": remote_path,
                "version": version,
                "source": source,
                "checksum_status": "remote_pending",
                "license_class": "open-data-with-terms" if source not in {"workbook", "provider_target_builder"} else "restricted-or-review",
                "build_required": build_required,
                "search_template": f"{engine}_provider_local_search",
                "retention_policy": "runpod_volume_persistent",
                "backup_policy": "external_backup_required",
                "priority": priority,
                "run_gate": run_gate,
                "cost_class": cost_class,
                "prep_roi": prep_roi,
                "bootstrap_strategy": bootstrap_strategy,
                "notes": "No database files are stored locally; provider volume only.",
            }
        )
    return rows


def build_cache_ledger() -> list[dict[str, str]]:
    records = [
        ("runpod_workspace", "network_volume_mount", "/workspace", "true", "500", "runpod_volume_persistent", "external_backup_required"),
        ("db_cache_root", "database_cache", DB_CACHE_ROOT, "true", "500", "runpod_volume_persistent", "external_backup_required"),
        ("search_cache", "search_result_cache", "/workspace/genecluster/search-cache", "true", "100", "runpod_volume_persistent", "external_backup_optional"),
        ("run_root", "run_root", "/workspace/genecluster/runs", "true", "200", "runpod_volume_persistent", "external_backup_required"),
        ("nextflow_cache", "nextflow_cache", "/workspace/genecluster/nextflow-cache", "true", "100", "runpod_volume_persistent", "external_backup_required"),
        ("sra_cache", "sra_cache", "/workspace/genecluster/sra-cache", "true", "200", "runpod_volume_persistent", "external_backup_required"),
        ("scratch", "fast_scratch", "/workspace/genecluster/scratch", "true", "200", "delete_after_review", "external_backup_optional"),
        ("summary_sync", "summary_export", f"{REMOTE_RUN_ROOT}/summary", "true", "10", "runpod_volume_persistent", "local_summary_sync_allowed"),
    ]
    rows = []
    for cache_id, role, remote_path, required, space, retention_policy, backup_policy in records:
        rows.append(
            {
                "cache_id": cache_id,
                "provider_class": "runpod_pod",
                "cache_role": role,
                "remote_path": remote_path,
                "mount_path": "/workspace",
                "required": required,
                "free_space_gb": space,
                "retention_policy": retention_policy,
                "backup_policy": backup_policy,
                "env_var": "GENECLUSTER_RUNPOD_NETWORK_VOLUME_ID" if cache_id == "runpod_workspace" else "",
                "notes": "Generated from Excel intake; no local heavy storage.",
            }
        )
    return rows


def build_pathway_steps(query_rows: list[dict[str, str]]) -> list[dict[str, str]]:
    buckets = [
        ("STEP_STR", "strictosidine formation", "anchor", "secologanin + tryptamine", "strictosidine", "strictosidine synthase"),
        ("STEP_SGD", "strictosidine aglycone formation", "anchor", "strictosidine", "strictosidine aglycone", "beta-glucosidase"),
        ("STEP_REDUCTASE", "post-strictosidine reductase branch", "candidate", "strictosidine aglycone-derived intermediates", "corynanthe-like reduced intermediates", "medium-chain dehydrogenase/reductase"),
        ("STEP_CYP", "oxidative tailoring / hydroxylation", "candidate", "corynanthe-like intermediates", "hydroxylated products", "cytochrome P450"),
        ("STEP_OMT", "methyltransferase tailoring", "candidate", "hydroxylated or acid/enol intermediates", "methylated corynanthe-like products", "methyltransferase"),
        ("STEP_TRANSPORT", "transport / regulatory neighborhood context", "context_only", "pathway metabolites", "transport or regulation context", "transporter"),
    ]
    rows: list[dict[str, str]] = []
    for step_id, step_name, role, substrate, product, enzyme in buckets:
        matching = [row["query_id"] for row in query_rows if row["enzyme_class"] == enzyme]
        if not matching:
            continue
        rows.append(
            {
                "pathway_step_id": step_id,
                "step_name": step_name,
                "role": role,
                "substrate": substrate,
                "product": product,
                "expected_enzyme_families": enzyme,
                "known_seed_examples": ";".join(row["query_name"] for row in query_rows if row["query_id"] in matching),
                "query_ids": ";".join(matching),
                "expected_evidence": "homology;domain;reciprocal;anchor;neighborhood_if_coordinates_available",
                "unresolved_caveats": "Product chemistry remains hypothesis without functional/metabolomics validation.",
                "claim_limit": "candidate_only" if role != "context_only" else "context_only",
            }
        )
    return rows


def write_query_fasta(path: Path, records: list[dict[str, str]]) -> None:
    lines: list[str] = []
    for record in records:
        lines.append(f">{record['query_id']} {record['name']}")
        seq = record["sequence"]
        lines.extend(seq[i : i + 80] for i in range(0, len(seq), 80))
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def build_campaign_manifest(
    campaign_id: str,
    data_rows: list[dict[str, str]],
    *,
    target_organism: str = "target species",
    target_common_name: str = "",
    target_pathway: str = "workbook-derived candidate discovery",
    target_scope: str = "Workbook-derived candidate search, genome anchoring, and neighborhood capture.",
) -> dict[str, Any]:
    accessions = [
        {"id": row["accession"], "role": row["data_role"], "source_url": row["source_url"]}
        for row in data_rows
    ]
    organism: dict[str, str] = {"scientific_name": target_organism}
    if target_common_name:
        organism["common_name"] = target_common_name
    return {
        "schema_version": 1,
        "campaign_id": campaign_id,
        "organism": organism,
        "target_pathway": {
            "name": target_pathway,
            "scope": target_scope,
            "caveat": "Product claims require external functional or metabolomics validation; spreadsheet-derived seeds require provenance review.",
        },
        "accessions": accessions,
        "query_set": {
            "ledger": "query-ledger.tsv",
            "sequence_fasta": "query-sequences.faa",
            "seed_classes": ["strictosidine synthase", "strictosidine glucosidase", "DCS/reductase", "CYP", "methyltransferase", "transporter_context"],
        },
        "run_readiness_contracts": {
            "project_goals": "project-goals.yaml",
            "pathway_steps": "pathway-steps.tsv",
            "database_ledger": "database-ledger.tsv",
            "cache_ledger": "cache-ledger.tsv",
            "database_tier": "maximum",
            "search_policy": "local_or_provider_blast_diamond_mmseqs_only",
            "return_policy": "summaries_manifests_ledgers_provenance_versions_claim_audit_dossier_only",
        },
        "evidence_classes": [
            "transcript_hit",
            "protein_hit",
            "domain_hit",
            "genome_localized",
            "neighborhood_supported",
            "coexpression_supported",
            "review_required",
        ],
        "execution": {
            "mode": "remote",
            "provider_class": "runpod_pod",
            "large_local_downloads": False,
            "remote_workdir": f"{REMOTE_RUN_ROOT}",
            "remote_volume_mount": "/workspace",
            "artifact_policy": "summaries_only",
            "web_tool_policy": "container-only",
            "future_provider_classes": ["local_lite", "local_full", "ssh_hpc", "cloud_vm", "managed_workflow"],
        },
        "run_scopes": {
            "smoke": {"description": "Metadata/query resolution and validation only.", "heavy_compute": False},
            "candidate_search": {"description": "Provider-side candidate search and small dossier.", "heavy_compute": True},
            "genome_context": {"description": "Genome anchoring and neighborhood capture with claim gates.", "heavy_compute": True},
            "full_public_mining": {"description": "Full public/approved-data mining with claim audit.", "heavy_compute": True},
        },
        "artifact_policy": {
            "local_downloads": "small_summaries_only",
            "remote_large_artifacts": "provider-volume-only",
            "forbidden_local_artifacts": ["fastq", "sra", "bam", "cram", "sam", "large_genome_assembly", "blast_database", "workflow_workdir"],
        },
        "license_policy": {
            "skill_repo_license_target": "MIT",
            "third_party_mode": "orchestrated_external_dependencies",
            "web_tool_policy": "container-only",
            "restricted_resources_require_approval": True,
        },
        "review_policy": {
            "statuses": ["new", "needs-human-review", "accepted", "rejected", "needs-rerun", "publication-candidate"],
            "cluster_claim_rule": "Physical gene-cluster claims require genome coordinates; transcript-only hits are candidate genes only.",
        },
    }


def build_project_goals(
    campaign_id: str,
    parsed: dict[str, Any],
    *,
    target_organism: str = "target species",
    target_pathway: str = "workbook-derived candidate discovery",
    default_run_scope: str = "candidate_search",
    provider_class: str = "runpod_pod",
) -> dict[str, Any]:
    goals = parsed["goals"] or [
        "Find all relevant transcripts.",
        "Find top matches to protein queries.",
        "Find clusters for selected queries.",
    ]
    return {
        "schema_version": 1,
        "project_id": campaign_id,
        "scientific_goal": " ".join(goals),
        "organism": {"scientific_name": target_organism, "outgroups": []},
        "target_pathway": target_pathway,
        "default_run_scope": default_run_scope,
        "database_tier": "maximum",
        "execution_defaults": {
            "provider_class": provider_class,
            "provider_role": "selected_provider",
            "mount_path": "/workspace",
            "remote_workdir_template": "/workspace/genecluster/runs/{run_id}",
            "db_cache_root": DB_CACHE_ROOT,
            "nextflow_cache_root": "/workspace/genecluster/nextflow-cache",
            "scratch_root": "/workspace/genecluster/scratch",
            "return_policy": "summaries_only",
        },
        "runtime_budget": {
            "target_runtime_hours": 24,
            "hard_stop_hours": 24,
            "completion_definition": "complete_summary_dossier_with_deferred_lane_manifest",
            "budget_policy": "finish_with_caveats_rather_than_extend_runtime",
            "defer_before_exceeding_budget": ["de_novo_transcriptome_assembly", "broad_nr_nt_searches", "full_interproscan", "coexpression_if_sample_design_insufficient"],
        },
        "priorities": {
            "novelty": "high",
            "deduplication": "high",
            "splice_variants": "medium",
            "genome_context": "high",
            "local_blast": "high",
            "coexpression": "medium",
            "synteny": "medium",
        },
        "allowed_compute_lanes": ["local_blast", "diamond", "mmseqs2", "hmmer", "target_db_build", "genome_neighborhood_capture", "orthology_synteny", "coexpression_if_supported"],
        "forbidden_compute_lanes": ["ncbi_remote_blast_batch", "public_webserver_private_upload", "local_repo_raw_download"],
        "stop_conditions": ["raw_or_heavy_local_artifact_detected", "provider_cache_preflight_failed", "missing_required_tool", "launch_manifest_claim_boundary_violation"],
        "claim_boundaries": {
            "transcriptome_cluster_claim": "forbidden",
            "broad_family_product_claim": "forbidden",
            "cluster_claim_requires": ["genome_coordinates", "neighboring_gene_evidence"],
            "product_claim_requires": ["functional_assay_or_lcms_validation", "validated_literature_seed_support"],
        },
        "approved_resources": {
            "data_ledgers": ["data-ledger.tsv"],
            "query_ledgers": ["query-ledger.tsv"],
            "resource_ledgers": ["resource-ledger.tsv"],
            "database_ledgers": ["database-ledger.tsv"],
        },
        "review_outputs": {
            "primary": "interactive_structured_dossier",
            "compatibility": "excel_export",
            "required_ledgers": ["candidate_hits.tsv", "evidence.jsonl", "provenance.jsonl", "claim-audit.jsonl"],
        },
    }


def intake_workbook(
    workbook: Path,
    out: Path,
    campaign_id: str,
    *,
    target_organism: str = "target species",
    target_common_name: str = "",
    target_pathway: str = "workbook-derived candidate discovery",
    target_scope: str = "Workbook-derived candidate search, genome anchoring, and neighborhood capture.",
    default_run_scope: str = "candidate_search",
    provider_class: str = "runpod_pod",
) -> Path:
    parsed = read_workbook(workbook)
    out.mkdir(parents=True, exist_ok=True)
    data_rows = build_data_ledger(parsed, campaign_id)
    query_rows, fasta_records = build_query_ledger(parsed)
    pathway_rows = build_pathway_steps(query_rows)

    data_columns = [
        "dataset_id",
        "accession",
        "run_id",
        "data_role",
        "sample_type",
        "organism",
        "bioproject",
        "technology",
        "expected_size",
        "source_url",
        "remote_path",
        "checksum_status",
        "data_sensitivity",
        "allowed_compute_location",
        "allowed_upload",
        "redistribution_policy",
        "terms_checked_date",
        "license_url",
        "citation_doi",
        "md5_or_sha256",
        "frozen_metadata_path",
        "raw_artifact_policy",
        "retention_policy",
        "operator_approval_id",
        "notes",
    ]
    query_columns = [
        "query_id",
        "query_name",
        "source_organism",
        "sequence_source",
        "enzyme_class",
        "pathway_role",
        "confidence",
        "citation",
        "resolved_accession",
        "sequence_type",
        "sequence_length",
        "checksum",
        "family_scope",
        "motif_requirements",
        "negative_controls",
        "decoy_or_broad_family_flag",
        "expected_false_positive_risk",
        "curation_status",
        "last_resolved_at",
        "notes",
    ]
    resource_columns = ["resource", "resource_type", "version", "license_class", "use_mode", "citation"]
    database_columns = [
        "db_id",
        "engine",
        "sequence_type",
        "remote_path",
        "version",
        "source",
        "checksum_status",
        "license_class",
        "build_required",
        "search_template",
        "retention_policy",
        "backup_policy",
        "priority",
        "run_gate",
        "cost_class",
        "prep_roi",
        "bootstrap_strategy",
        "notes",
    ]
    cache_columns = [
        "cache_id",
        "provider_class",
        "cache_role",
        "remote_path",
        "mount_path",
        "required",
        "free_space_gb",
        "retention_policy",
        "backup_policy",
        "env_var",
        "notes",
    ]
    pathway_columns = [
        "pathway_step_id",
        "step_name",
        "role",
        "substrate",
        "product",
        "expected_enzyme_families",
        "known_seed_examples",
        "query_ids",
        "expected_evidence",
        "unresolved_caveats",
        "claim_limit",
    ]

    write_json(
        out / "campaign-manifest.json",
        build_campaign_manifest(
            campaign_id,
            data_rows,
            target_organism=target_organism,
            target_common_name=target_common_name,
            target_pathway=target_pathway,
            target_scope=target_scope,
        ),
    )
    write_json(
        out / "project-goals.yaml",
        build_project_goals(
            campaign_id,
            parsed,
            target_organism=target_organism,
            target_pathway=target_pathway,
            default_run_scope=default_run_scope,
            provider_class=provider_class,
        ),
    )
    write_tsv(out / "data-ledger.tsv", data_columns, data_rows)
    write_tsv(out / "query-ledger.tsv", query_columns, query_rows)
    write_tsv(out / "resource-ledger.tsv", resource_columns, build_resource_ledger())
    write_tsv(out / "database-ledger.tsv", database_columns, build_database_ledger())
    write_tsv(out / "cache-ledger.tsv", cache_columns, build_cache_ledger())
    write_tsv(out / "pathway-steps.tsv", pathway_columns, pathway_rows)
    write_query_fasta(out / "query-sequences.faa", fasta_records)
    write_json(
        out / "intake-report.json",
        {
            "schema_version": 1,
            "source_workbook": str(workbook),
            "campaign_id": campaign_id,
            "target_organism": target_organism,
            "target_pathway": target_pathway,
            "default_run_scope": default_run_scope,
            "provider_class": provider_class,
            "sheet": parsed["sheet"],
            "datasets": len(data_rows),
            "queries": len(query_rows),
            "embedded_query_sequences": len(fasta_records),
            "resolver_required_queries": [row["query_id"] for row in query_rows if row["curation_status"] == "remote_resolve_required"],
            "goals": parsed["goals"],
            "local_artifact_policy": "small control-plane files only; no raw sequence downloads",
        },
    )
    return out / "campaign-manifest.json"


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", required=True, type=Path, help="Old-style GeneCluster request workbook (.xlsx).")
    parser.add_argument("--out", required=True, type=Path, help="Output directory for generated campaign ledgers.")
    parser.add_argument("--campaign-id", default="genecluster-workbook-intake-v0")
    parser.add_argument("--target-organism", default="target species")
    parser.add_argument("--target-common-name", default="")
    parser.add_argument("--target-pathway", default="workbook-derived candidate discovery")
    parser.add_argument("--target-scope", default="Workbook-derived candidate search, genome anchoring, and neighborhood capture.")
    parser.add_argument("--default-run-scope", default="candidate_search")
    parser.add_argument("--provider-class", default="runpod_pod")
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    manifest = intake_workbook(
        args.workbook,
        args.out,
        args.campaign_id,
        target_organism=args.target_organism,
        target_common_name=args.target_common_name,
        target_pathway=args.target_pathway,
        target_scope=args.target_scope,
        default_run_scope=args.default_run_scope,
        provider_class=args.provider_class,
    )
    print(f"Wrote GeneCluster Excel intake bundle: {args.out}")
    print(f"Campaign manifest: {manifest}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
